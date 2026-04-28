"""Tree Navigator service — extracts cells from AssumptionPack + Run output files
into flat lists for the tree's right-pane Inputs/Outputs detail views.

Endpoints exposed via routers/tree.py:
  GET /api/projects/{pid}/assumption-packs/{sid}/inputs
      -> [{tab, cell_ref, label, value, type}]
  GET /api/projects/{pid}/assumption-packs/{sid}/outputs
      -> [{tab, cell_ref, label, value, run_id, model_id, model_version,
           output_template_id, output_template_version}]
  GET /api/projects/{pid}/assumption-packs/{sid}/inputs/{tab}/{cell_ref}
      -> single-cell focus view
  GET /api/projects/{pid}/assumption-packs/{sid}/outputs/{tab}/{cell_ref}/history
      -> [{run_id, value, started_at}]  for the time-series chart
"""
from __future__ import annotations

import io
import os
import tempfile
from typing import Any

import openpyxl
from openpyxl.cell.cell import MergedCell


def _detect_label(ws, row: int, col: int) -> str | None:
    """Heuristic: look at the cell to the LEFT of (row, col); if it's a string, use it.
    Otherwise look at the cell ABOVE; if it's a string, use it. Else return None."""
    if col > 1:
        left = ws.cell(row=row, column=col - 1).value
        if isinstance(left, str) and left.strip():
            return left.strip()
    if row > 1:
        above = ws.cell(row=row - 1, column=col).value
        if isinstance(above, str) and above.strip():
            return above.strip()
    return None


def _detect_type(value: Any, number_format: str | None = None) -> str:
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float)):
        if number_format and "%" in number_format:
            return "percentage"
        if number_format and any(s in number_format for s in ("$", "€", "£", "¥")):
            return "currency"
        return "number"
    if isinstance(value, str):
        return "text"
    # datetime, date, time
    cls = type(value).__name__.lower()
    if "date" in cls or "time" in cls:
        return "date"
    return "unknown"


def list_input_cells(xlsx_bytes: bytes) -> list[dict[str, Any]]:
    """Read an AssumptionPack file, return a flat list of all populated I_ cells.

    Each cell entry: {tab, cell_ref, row, column, label, value, type}.
    Skips cells whose value is None or a formula (formulas show #REF! standalone).
    Sorted by tab order then by row, then column.
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    out: list[dict[str, Any]] = []
    try:
        for tab in wb.sheetnames:
            if not tab.startswith("I_"):
                continue
            ws = wb[tab]
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c, MergedCell):
                        continue
                    if c.value is None:
                        continue
                    if isinstance(c.value, str) and c.value.startswith("="):
                        continue
                    label = _detect_label(ws, c.row, c.column)
                    out.append(
                        {
                            "tab": tab,
                            "cell_ref": c.coordinate,
                            "row": c.row,
                            "column": c.column,
                            "label": label,
                            "value": c.value,
                            "type": _detect_type(c.value, c.number_format),
                        }
                    )
        return out
    finally:
        wb.close()


def list_output_cells(output_xlsx_bytes: bytes) -> list[dict[str, Any]]:
    """Read a Run's output .xlsx (which is a calculated OutputTemplate workbook)
    and return its O_ cell values.

    Each cell entry: {tab, cell_ref, label, value, type}.
    """
    # We have to write to a temp file because openpyxl needs a path for data_only=True
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(output_xlsx_bytes)
        tmp_path = tmp.name
    try:
        wb = openpyxl.load_workbook(tmp_path, data_only=True)
    finally:
        os.unlink(tmp_path)

    out: list[dict[str, Any]] = []
    try:
        for tab in wb.sheetnames:
            if not tab.startswith("O_"):
                continue
            ws = wb[tab]
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c, MergedCell):
                        continue
                    if c.value is None:
                        continue
                    label = _detect_label(ws, c.row, c.column)
                    out.append(
                        {
                            "tab": tab,
                            "cell_ref": c.coordinate,
                            "row": c.row,
                            "column": c.column,
                            "label": label,
                            "value": c.value,
                            "type": _detect_type(c.value, c.number_format),
                        }
                    )
        return out
    finally:
        wb.close()
