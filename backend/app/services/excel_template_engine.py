"""Excel Template engine — tab classification, Scenario extraction, overlay, validate.

The architecture in one line:
  Template = .xlsx with tabs prefixed I_ / O_ / other (calc)
  Scenario = .xlsx containing only the Template's I_ tabs (edited by humans)
  Calculate = cell-copy Scenario's I_ tabs over Template's I_ tabs, recalc.

Prefix matching is CASE-SENSITIVE. "I_Rev" is an input; "i_Rev" is a calc tab.

Spike proven on Campus_Adele_Model_20260416_1410.xlsx — see /tmp/mfm_spike_*.py.
"""
from __future__ import annotations

import io
import logging
from copy import copy
from typing import Any

import openpyxl
from openpyxl.cell.cell import MergedCell

logger = logging.getLogger(__name__)

INPUT_PREFIX = "I_"
OUTPUT_PREFIX = "O_"
M_PREFIX = "M_"

# Roles for validate_template — used by Sprint A and beyond.
ROLE_MODEL = "model"
ROLE_ASSUMPTION_PACK = "assumption_pack"
ROLE_OUTPUT_TEMPLATE_XLSX = "output_template_xlsx"


def classify_tabs(wb: openpyxl.Workbook) -> dict[str, list[str]]:
    """Classify a workbook's tabs by case-sensitive prefix.

    Returns {"input_tabs", "output_tabs", "m_tabs", "calc_tabs"}.
    Ordering follows the sheet order in the workbook.
    """
    input_tabs: list[str] = []
    output_tabs: list[str] = []
    m_tabs: list[str] = []
    calc_tabs: list[str] = []
    for name in wb.sheetnames:
        if name.startswith(INPUT_PREFIX):
            input_tabs.append(name)
        elif name.startswith(OUTPUT_PREFIX):
            output_tabs.append(name)
        elif name.startswith(M_PREFIX):
            m_tabs.append(name)
        else:
            calc_tabs.append(name)
    return {
        "input_tabs": input_tabs,
        "output_tabs": output_tabs,
        "m_tabs": m_tabs,
        "calc_tabs": calc_tabs,
    }


def classify_bytes(content: bytes) -> dict[str, list[str]]:
    """Convenience: classify tabs directly from file bytes."""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
    try:
        return classify_tabs(wb)
    finally:
        wb.close()


def validate_template(wb: openpyxl.Workbook, role: str = ROLE_MODEL) -> list[str]:
    """Return a list of human-readable validation errors. Empty list = valid.

    Universal rules (all roles):
      - No duplicate tab names (openpyxl enforces this but we surface a clearer error).
      - Tab names must not begin with an ASCII dot (Excel rejects these).

    Per-role rules:
      ROLE_MODEL                 — must have ≥1 I_; must have ≥1 O_; must NOT have M_.
      ROLE_ASSUMPTION_PACK       — must have ≥1 I_; must NOT have O_, M_, or calc tabs.
      ROLE_OUTPUT_TEMPLATE_XLSX  — must have ≥1 O_; must NOT have I_; M_ optional but if any
                                    must match a Model O_ at run-validate time (separate check).
    """
    errors: list[str] = []
    if len(wb.sheetnames) != len(set(wb.sheetnames)):
        errors.append("Workbook has duplicate tab names.")
    for n in wb.sheetnames:
        if n.startswith("."):
            errors.append(f"Tab name {n!r} is invalid (starts with a dot).")

    classes = classify_tabs(wb)

    if role == ROLE_MODEL:
        if not classes["input_tabs"]:
            errors.append("Model has no I_ tabs. At least one input tab is required.")
        if not classes["output_tabs"]:
            errors.append("Model has no O_ tabs. At least one output tab is required.")
        if classes["m_tabs"]:
            errors.append(
                "Model contains M_ tabs (only OutputTemplates may have M_ tabs): "
                + ", ".join(classes["m_tabs"])
            )
    elif role == ROLE_ASSUMPTION_PACK:
        if not classes["input_tabs"]:
            errors.append("AssumptionPack has no I_ tabs.")
        non_input = classes["output_tabs"] + classes["m_tabs"] + classes["calc_tabs"]
        if non_input:
            errors.append(
                "AssumptionPack contains tabs that are not I_ tabs: "
                + ", ".join(non_input)
            )
    elif role == ROLE_OUTPUT_TEMPLATE_XLSX:
        if not classes["output_tabs"]:
            errors.append("OutputTemplate has no O_ tabs (the user-facing artifact).")
        if classes["input_tabs"]:
            errors.append(
                "OutputTemplate contains I_ tabs (those belong on Models/AssumptionPacks): "
                + ", ".join(classes["input_tabs"])
            )
    else:
        errors.append(f"Unknown role: {role!r}")

    return errors


def basename(tab_name: str) -> str:
    """Strip the I_/O_/M_ prefix to get the matching key. 'I_Foo' -> 'Foo'."""
    for p in (INPUT_PREFIX, OUTPUT_PREFIX, M_PREFIX):
        if tab_name.startswith(p):
            return tab_name.removeprefix(p)
    return tab_name


def extract_model_outputs(recalced_model_bytes: bytes) -> dict[str, dict[str, Any]]:
    """Read all O_* tab cells from a recalculated Model workbook.

    Returns {"O_<tab>": {"<cell_ref>": <value>}}.
    Uses data_only=True to read cached values produced by LibreOffice recalc.
    """
    import os
    import tempfile

    # openpyxl needs a real path for data_only — use a temp file
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(recalced_model_bytes)
        tmp_path = tmp.name
    try:
        wb = openpyxl.load_workbook(tmp_path, data_only=True)
    finally:
        os.unlink(tmp_path)

    out: dict[str, dict[str, Any]] = {}
    try:
        for tab in wb.sheetnames:
            if not tab.startswith(OUTPUT_PREFIX):
                continue
            ws = wb[tab]
            cells: dict[str, Any] = {}
            for row in ws.iter_rows():
                for c in row:
                    if c.value is not None and not isinstance(c, MergedCell):
                        cells[c.coordinate] = c.value
            out[tab] = cells
        return out
    finally:
        wb.close()


def overlay_outputs_onto_template(
    output_template_bytes: bytes,
    model_outputs: dict[str, dict[str, Any]],
) -> tuple[bytes, list[str]]:
    """Inject Model O_<name> values into OutputTemplate M_<name> cells.

    Returns (merged_template_bytes, warnings).
    Skips M_ tabs that have no matching O_ tab (validator should have caught this;
    surfaces a warning here as defense in depth).
    """
    wb = openpyxl.load_workbook(io.BytesIO(output_template_bytes), data_only=False)
    warnings: list[str] = []
    try:
        for tab_name in wb.sheetnames:
            if not tab_name.startswith(M_PREFIX):
                continue
            base = tab_name.removeprefix(M_PREFIX)
            source_o_tab = OUTPUT_PREFIX + base
            if source_o_tab not in model_outputs:
                warnings.append(
                    f"OutputTemplate's {tab_name!r} has no matching {source_o_tab!r} in Model output"
                )
                continue

            dst_ws = wb[tab_name]
            cell_values = model_outputs[source_o_tab]

            # Unmerge destination first so we can write everywhere
            for merge in list(dst_ws.merged_cells.ranges):
                dst_ws.unmerge_cells(str(merge))

            for cell_ref, value in cell_values.items():
                try:
                    dst_ws[cell_ref].value = value
                except (AttributeError, ValueError) as e:
                    warnings.append(f"Could not write {cell_ref} on {tab_name!r}: {e}")

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue(), warnings
    finally:
        wb.close()


def extract_scenario_from_template(template_bytes: bytes) -> bytes:
    """Produce an inputs-only .xlsx from a Template — the initial Scenario file.

    Deletes every tab that isn't I_*. The resulting file is what the user edits
    in Drive / Excel. Note: some cells in I_ tabs may reference calc tabs and
    will display #REF! when opened standalone; they resolve correctly once
    overlaid back into the Template for calculation.
    """
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes), data_only=False)
    try:
        to_delete = [n for n in wb.sheetnames if not n.startswith(INPUT_PREFIX)]
        for n in to_delete:
            del wb[n]
        if not wb.sheetnames:
            # openpyxl requires at least one sheet; add a placeholder if somehow empty
            wb.create_sheet("I_Empty")
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    finally:
        wb.close()


def _overlay_tab(src_ws, dst_ws) -> None:
    """Cell-by-cell copy of src_ws content into dst_ws, including styles and merges.

    Destination merged ranges are unmerged first (MergedCells are read-only),
    then destination cells within the source's used range are cleared, then
    source cells are copied, then source merges are re-applied.
    """
    # Step 1: unmerge destination
    for merge in list(dst_ws.merged_cells.ranges):
        dst_ws.unmerge_cells(str(merge))

    # Step 2: clear destination cells within the source's used range, to guard
    # against stale rows/columns left over from a previous (larger) version.
    max_row = src_ws.max_row or 1
    max_col = src_ws.max_column or 1
    for row in dst_ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for c in row:
            c.value = None

    # Step 3: copy source cells (value + style)
    for row in src_ws.iter_rows():
        for sc in row:
            if isinstance(sc, MergedCell):
                continue
            dc = dst_ws.cell(row=sc.row, column=sc.column)
            dc.value = sc.value
            if sc.has_style:
                dc.font = copy(sc.font)
                dc.fill = copy(sc.fill)
                dc.border = copy(sc.border)
                dc.alignment = copy(sc.alignment)
                dc.number_format = sc.number_format
                dc.protection = copy(sc.protection)

    # Step 4: column widths and row heights
    for col_letter, dim in src_ws.column_dimensions.items():
        if dim.width:
            dst_ws.column_dimensions[col_letter].width = dim.width
    for r, dim in src_ws.row_dimensions.items():
        if dim.height:
            dst_ws.row_dimensions[r].height = dim.height

    # Step 5: re-apply source merges
    for merge in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merge))


def overlay_scenario_on_template(
    template_bytes: bytes, scenario_bytes: bytes
) -> tuple[bytes, list[str]]:
    """Merge a Scenario's I_ tabs into a copy of the Template workbook.

    Returns (merged_file_bytes, warnings).
    Raises ValueError on hard contract violations.
    """
    tpl = openpyxl.load_workbook(io.BytesIO(template_bytes), data_only=False)
    sc = openpyxl.load_workbook(io.BytesIO(scenario_bytes), data_only=False)
    warnings: list[str] = []
    try:
        tpl_classes = classify_tabs(tpl)
        sc_classes = classify_tabs(sc)

        # Contract: Scenario must contain ONLY I_ tabs
        if sc_classes["output_tabs"] or sc_classes["calc_tabs"]:
            bad = sc_classes["output_tabs"] + sc_classes["calc_tabs"]
            raise ValueError(
                "Scenario file contains tabs that are not I_ tabs: "
                + ", ".join(bad)
            )

        # Contract: every I_ tab the Template declares must be in the Scenario
        tpl_inputs = set(tpl_classes["input_tabs"])
        sc_inputs = set(sc_classes["input_tabs"])
        missing = tpl_inputs - sc_inputs
        if missing:
            raise ValueError(
                "Scenario is missing required input tabs: " + ", ".join(sorted(missing))
            )

        # Warn if Scenario has extra I_ tabs the Template doesn't know about
        extra = sc_inputs - tpl_inputs
        if extra:
            warnings.append(
                "Scenario has extra I_ tabs not in Template (ignored): "
                + ", ".join(sorted(extra))
            )

        # Overlay each I_ tab
        for name in tpl_classes["input_tabs"]:
            _overlay_tab(sc[name], tpl[name])

        buf = io.BytesIO()
        tpl.save(buf)
        return buf.getvalue(), warnings
    finally:
        tpl.close()
        sc.close()


def replace_template_tabs(
    existing_template_bytes: bytes, new_template_bytes: bytes
) -> tuple[bytes, dict[str, Any]]:
    """Upgrade an existing Template with I_/O_ tabs from a new upload (Option A).

    Behavior per agreed design:
      - Calc tabs (non-prefixed): replaced wholesale from the new file.
      - I_ tabs: each I_ tab in the new file replaces the corresponding one in
        the existing template. New I_ tabs are added. I_ tabs only in the old
        template are removed.
      - O_ tabs: same as I_ tabs (derived, no user data).

    In practice this is identical to "the new upload becomes the Template",
    but named distinctly so routers can express intent. The function also
    reports a diff so the caller can surface a confirmation to the user.
    """
    # Classify both for the diff report
    old_classes = classify_bytes(existing_template_bytes)
    new_classes = classify_bytes(new_template_bytes)

    def diff(old: list[str], new: list[str]) -> dict[str, list[str]]:
        old_set, new_set = set(old), set(new)
        return {
            "added": sorted(new_set - old_set),
            "removed": sorted(old_set - new_set),
            "kept": sorted(old_set & new_set),
        }

    report = {
        "input_tabs": diff(old_classes["input_tabs"], new_classes["input_tabs"]),
        "output_tabs": diff(old_classes["output_tabs"], new_classes["output_tabs"]),
        "calc_tabs": diff(old_classes["calc_tabs"], new_classes["calc_tabs"]),
    }
    # Option A: the new file IS the new Template content.
    return new_template_bytes, report


def calculate(
    template_bytes: bytes, scenario_bytes: bytes
) -> dict[str, Any]:
    """Full Calculate pipeline: overlay + LibreOffice recalc.

    Returns {
      "output_bytes": bytes | None,
      "merged_bytes": bytes,        # pre-recalc, in case caller wants it
      "warnings": [...],
      "recalculated": bool,
    }
    """
    from backend.app.services import excel_engine

    merged_bytes, warnings = overlay_scenario_on_template(template_bytes, scenario_bytes)
    recalced = excel_engine.recalculate_with_libreoffice(merged_bytes)
    return {
        "output_bytes": recalced or merged_bytes,
        "merged_bytes": merged_bytes,
        "warnings": warnings,
        "recalculated": recalced is not None,
    }
