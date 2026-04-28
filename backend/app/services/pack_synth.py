"""Sprint I — synthesize pack xlsx bytes from a cell-overrides dict.

Lets json packs and pull packs reuse the existing Run engine unchanged.
The engine path is:

    pack_bytes  ─┐
                 │
    model_bytes ─┼─→ run_executor.execute_run_sync()
                 │
    tpl_bytes   ─┘

For json/pull packs there's no real pack file. We build an in-memory xlsx
with one tab per ``I_*`` entry in the overrides dict, then return its bytes
so the engine can consume it identically to an xlsx-kind pack.

Non-numeric values are written as-is (openpyxl handles strings, dates,
bools, etc.).
"""
from __future__ import annotations

import io
from typing import Any

import openpyxl


def synthesize_pack_xlsx_from_overrides(
    overrides: dict[str, dict[str, Any]],
) -> bytes:
    """Build an in-memory xlsx with I_* tabs filled from ``overrides``.

    overrides shape::

        {
            "I_Numbers":  {"B1": 5, "B2": 7},
            "I_Pricing":  {"B5": 1500, "B6": 2000},
        }

    Tabs whose names don't start with ``I_`` are still written (the merge
    code only looks at I_*, but other tabs are harmless and can help with
    debugging).
    """
    wb = openpyxl.Workbook()
    # openpyxl auto-creates "Sheet" — drop it so the ordering of our tabs is
    # natural (purely the keys' iteration order).
    if wb.active.title in wb.sheetnames:
        wb.remove(wb.active)

    for tab, cells in overrides.items():
        ws = wb.create_sheet(tab)
        for cell_ref, value in cells.items():
            try:
                ws[cell_ref] = value
            except Exception:  # noqa: BLE001 — bad cell refs shouldn't fail the run
                # Skip silently; calling code may already have warned.
                continue

    # If overrides was empty we still need at least one sheet for the engine
    # not to choke. Add an empty placeholder.
    if not wb.sheetnames:
        wb.create_sheet("I_Empty")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
