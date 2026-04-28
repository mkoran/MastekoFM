"""Excel calculation engine — inject values, recalculate, extract results.

Flow: openpyxl injects → save temp .xlsx → LibreOffice recalculates → openpyxl reads results.
"""
import io
import logging
import os
import shutil
import subprocess
import tempfile
from typing import Any

import openpyxl

logger = logging.getLogger(__name__)

LIBREOFFICE_TIMEOUT = 120  # seconds per step (two conversions)
# Bumped from 60s → 120s after observing cold-start LO conversions on Cloud
# Run hitting the limit even on Hello World. Empirical: ~85s per Hello World
# run end-to-end (two stages + Drive ops). 120s gives headroom per stage.


def _find_libreoffice() -> str | None:
    """Find LibreOffice binary."""
    for path in ["/usr/bin/libreoffice", "/usr/bin/soffice", shutil.which("libreoffice"), shutil.which("soffice")]:
        if path and os.path.isfile(path):
            return path
    return None


def inject_values(wb: openpyxl.Workbook, sheet_name: str, cell_map: dict[str, Any]) -> int:
    """Inject values into specific cells of a worksheet.

    cell_map: {"B5": "Campus Adele", "B7": "2025-03-01", "B8": 13, ...}
    Skips merged cells gracefully. Returns count of successfully injected values.
    """
    ws = wb[sheet_name]
    injected = 0
    for cell_ref, value in cell_map.items():
        try:
            cell = ws[cell_ref]
            if hasattr(cell, "value") and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = value
                injected += 1
            else:
                logger.warning("Skipping merged/read-only cell %s in %s", cell_ref, sheet_name)
        except (AttributeError, TypeError):
            logger.warning("Could not inject into cell %s in %s", cell_ref, sheet_name)
    return injected


def inject_table_data(wb: openpyxl.Workbook, sheet_name: str, start_row: int, columns: list[str], rows: list[dict[str, Any]]) -> None:
    """Inject table data into a worksheet starting at a specific row.

    columns: ["B", "C", "D", ...] — column letters
    rows: list of dicts with column values
    """
    ws = wb[sheet_name]
    for i, row_data in enumerate(rows):
        for col_letter, key in zip(columns, row_data.keys(), strict=False):
            cell_ref = f"{col_letter}{start_row + i}"
            ws[cell_ref] = row_data[key]


def xlsx_to_pdf(file_bytes: bytes) -> bytes | None:
    """Convert .xlsx bytes to PDF bytes via LibreOffice headless.

    Sprint D-1: Used to publish a PDF artifact alongside the xlsx output of
    every Run whose OutputTemplate has ``pdf_export_xlsx = True``. The PDF
    is whatever LibreOffice would render for the workbook's print/page-setup
    — same fidelity Marc gets when he File → Download → PDF in Sheets.

    Returns the PDF bytes, or ``None`` if LibreOffice is unavailable or the
    conversion produced no output (best-effort; callers should treat None as
    "skip this artifact, don't fail the run").
    """
    lo_path = _find_libreoffice()
    if not lo_path:
        logger.warning("LibreOffice not found — skipping xlsx→pdf conversion")
        return None

    with tempfile.TemporaryDirectory() as tmpdir:
        input_path = os.path.join(tmpdir, "input.xlsx")
        with open(input_path, "wb") as f:
            f.write(file_bytes)

        try:
            env = {**os.environ, "HOME": tmpdir}
            r = subprocess.run(
                [lo_path, "--headless", "--norestore", "--calc", "--convert-to", "pdf",
                 "--outdir", tmpdir, input_path],
                capture_output=True, timeout=LIBREOFFICE_TIMEOUT, cwd=tmpdir, env=env,
            )
            logger.info("xlsx→pdf: exit=%d", r.returncode)

            pdf_path = os.path.join(tmpdir, "input.pdf")
            if not os.path.exists(pdf_path):
                # Fallback: scan dir for any .pdf
                for fname in os.listdir(tmpdir):
                    if fname.endswith(".pdf"):
                        pdf_path = os.path.join(tmpdir, fname)
                        break
                else:
                    logger.warning(
                        "xlsx→pdf produced no output. stderr=%s",
                        (r.stderr or b"").decode("utf-8", errors="replace")[:500],
                    )
                    return None

            with open(pdf_path, "rb") as f:
                return f.read()
        except subprocess.TimeoutExpired:
            logger.warning("xlsx→pdf timed out after %ds", LIBREOFFICE_TIMEOUT)
            return None


def recalculate_with_libreoffice(file_bytes: bytes) -> bytes | None:
    """Recalculate an .xlsx file using LibreOffice headless.

    Returns the recalculated file bytes, or None if LibreOffice is not available.
    """
    lo_path = _find_libreoffice()
    if not lo_path:
        logger.warning("LibreOffice not found — skipping recalculation")
        return None

    with tempfile.TemporaryDirectory() as tmpdir:
        input_path = os.path.join(tmpdir, "input.xlsx")
        with open(input_path, "wb") as f:
            f.write(file_bytes)

        try:
            env = {**os.environ, "HOME": tmpdir}

            # Step 1: Convert xlsx → ods (forces LibreOffice to fully parse and recalculate)
            r1 = subprocess.run(
                [lo_path, "--headless", "--norestore", "--calc", "--convert-to", "ods", "--outdir", tmpdir, input_path],
                capture_output=True, timeout=LIBREOFFICE_TIMEOUT, cwd=tmpdir, env=env,
            )
            logger.info("xlsx→ods: exit=%d", r1.returncode)

            ods_path = os.path.join(tmpdir, "input.ods")
            if not os.path.exists(ods_path):
                logger.error("ODS not produced. Dir: %s", os.listdir(tmpdir))
                return None

            # Step 2: Convert ods → xlsx (writes back with recalculated values cached)
            r2 = subprocess.run(
                [lo_path, "--headless", "--norestore", "--calc", "--convert-to", "xlsx", "--outdir", tmpdir, ods_path],
                capture_output=True, timeout=LIBREOFFICE_TIMEOUT, cwd=tmpdir, env=env,
            )
            logger.info("ods→xlsx: exit=%d", r2.returncode)

            # The output is input.xlsx (from input.ods converted)
            final_path = os.path.join(tmpdir, "input.xlsx")
            if not os.path.exists(final_path):
                # Check for alternative names
                for fname in os.listdir(tmpdir):
                    if fname.endswith(".xlsx"):
                        final_path = os.path.join(tmpdir, fname)
                        break

            if os.path.exists(final_path):
                with open(final_path, "rb") as f:
                    result_bytes = f.read()
                logger.info("LibreOffice produced %d bytes (input was %d)", len(result_bytes), len(file_bytes))
                return result_bytes

            logger.error("Final xlsx not found. Dir: %s", os.listdir(tmpdir))
            return None

        except subprocess.TimeoutExpired:
            logger.error("LibreOffice timed out after %ds", LIBREOFFICE_TIMEOUT)
            return None
        except Exception:
            logger.exception("LibreOffice recalculation failed")
            return None


def extract_cell_values(wb: openpyxl.Workbook, extractions: dict[str, list[str]]) -> dict[str, dict[str, Any]]:
    """Extract values from specific cells across sheets.

    extractions: {"Annual Summary": ["B9", "C9", "D9", ...], "Sources & Uses": ["B6", "B7"]}
    Returns: {"Annual Summary": {"B9": 121408, "C9": 1486034, ...}}
    """
    results: dict[str, dict[str, Any]] = {}
    for sheet_name, cells in extractions.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        sheet_results: dict[str, Any] = {}
        for cell_ref in cells:
            sheet_results[cell_ref] = ws[cell_ref].value
        results[sheet_name] = sheet_results
    return results


def extract_table(wb: openpyxl.Workbook, sheet_name: str, start_row: int, end_row: int, columns: dict[str, str]) -> list[dict[str, Any]]:
    """Extract a table range from a worksheet.

    columns: {"A": "label", "B": "year_1", "C": "year_2", ...}
    Returns list of row dicts.
    """
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = []
    for row_idx in range(start_row, end_row + 1):
        row_data: dict[str, Any] = {}
        has_data = False
        for col_letter, key in columns.items():
            val = ws[f"{col_letter}{row_idx}"].value
            row_data[key] = val
            if val is not None:
                has_data = True
        if has_data:
            rows.append(row_data)
    return rows


def calculate_model(model_bytes: bytes, assumptions: dict[str, Any]) -> dict[str, Any]:
    """Full calculation pipeline for a financial model.

    1. Load the .xlsx template
    2. Inject assumption values into input cells
    3. Recalculate with LibreOffice (or return raw if unavailable)
    4. Extract output values

    Returns dict with all calculated outputs.
    """
    # Step 1: Try injection + LibreOffice recalculation
    recalced_bytes = None
    try:
        wb = openpyxl.load_workbook(io.BytesIO(model_bytes))
        kv_map = assumptions.get("key_values", {})
        if kv_map and "Inputs & Assumptions" in wb.sheetnames:
            count = inject_values(wb, "Inputs & Assumptions", kv_map)
            logger.info("Injected %d/%d values", count, len(kv_map))
        for table_inject in assumptions.get("table_injections", []):
            inject_table_data(wb, table_inject["sheet"], table_inject["start_row"],
                              table_inject["columns"], table_inject["rows"])
        buf = io.BytesIO()
        wb.save(buf)
        wb.close()
        recalced_bytes = recalculate_with_libreoffice(buf.getvalue())
    except Exception:
        logger.exception("Injection/recalc failed")

    # Step 2: Read results — write to temp file for reliable data_only parsing
    source = recalced_bytes if recalced_bytes else model_bytes
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(source)
        tmp_path = tmp.name
    try:
        result_wb = openpyxl.load_workbook(tmp_path, data_only=True)
    finally:
        os.unlink(tmp_path)

    # Extract outputs
    outputs: dict[str, Any] = {}

    # Sources & Uses
    outputs["sources_and_uses"] = extract_table(
        result_wb, "Sources & Uses - Construction", 4, 30,
        {"A": "item", "B": "amount", "C": "notes"}
    )

    # Annual Summary
    outputs["annual_summary"] = extract_table(
        result_wb, "Annual Summary", 3, 130,
        {"A": "item", "B": "year_1", "C": "year_2", "D": "year_3", "E": "year_4", "F": "year_5", "G": "year_6", "H": "total", "I": "disposition"}
    )

    # Construction Budget Summary
    outputs["budget_summary"] = extract_table(
        result_wb, "Construction Budget & Draws", 4, 17,
        {"A": "category", "B": "amount"}
    )

    # Senior Construction Financing params
    if "Senior Construction Financing" in result_wb.sheetnames:
        sf = result_wb["Senior Construction Financing"]
        outputs["construction_loan"] = {
            "ltc": sf["B5"].value,
            "interest_rate": sf["B6"].value,
            "commitment_fee": sf["B7"].value,
            "total_project_cost": sf["B8"].value,
            "max_loan_amount": sf["B9"].value,
        }

    # Permanent Financing params
    if "Permanent Financing" in result_wb.sheetnames:
        pf = result_wb["Permanent Financing"]
        outputs["permanent_loan"] = {
            "loan_amount": pf["B4"].value,
            "interest_rate": pf["B5"].value,
            "amortization_years": pf["B7"].value,
            "loan_term_months": pf["B9"].value,
            "monthly_payment": pf["B10"].value,
        }

    result_wb.close()
    return outputs
