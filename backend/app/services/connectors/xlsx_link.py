"""Sprint I-1 — XLSX-Link connector.

Reads cells from another .xlsx file in the user's Drive. Uses the existing
Drive token (drive.file scope) — no new OAuth, no new secrets.

Three query shapes:

    Cell mode:   target="I_Numbers.B1"
                 config={"drive_file_id": "...", "sheet": "Master", "cell": "B1"}

    Range mode:  target="I_Numbers.B1"   (single value via aggregation)
                 config={"drive_file_id": "...", "sheet": "Master",
                         "range": "B1:B20", "aggregate": "sum"}

    Tab mode:    target="I_Numbers"
                 config={"drive_file_id": "...", "sheet": "Master"}
                 → every non-empty cell on Master is overlaid onto I_Numbers
                   at the same coordinate.

The Drive file is downloaded once per query (no cross-query cache yet —
cache is Sprint I-4). Empty cells are skipped on tab overlays (so a sparse
source doesn't blow away cells the model already had).
"""
from __future__ import annotations

import datetime as _dt
import io
import logging
from typing import Any

import openpyxl

from backend.app.models.assumption_pack import PullQuery
from backend.app.services import drive_service

logger = logging.getLogger(__name__)


def _download_xlsx(drive_file_id: str, user_token: str | None) -> bytes:
    bytes_ = drive_service.download_file(drive_file_id, user_access_token=user_token)
    if bytes_ is None:
        raise RuntimeError(f"Drive download failed for file {drive_file_id}")
    return bytes_


def _aggregate(values: list[Any], op: str) -> Any:
    """sum | mean | count | min | max — over numeric values only."""
    nums = [v for v in values if isinstance(v, int | float) and not isinstance(v, bool)]
    if op == "count":
        return len([v for v in values if v is not None])
    if not nums:
        return None
    if op == "sum":
        return sum(nums)
    if op == "mean":
        return sum(nums) / len(nums)
    if op == "min":
        return min(nums)
    if op == "max":
        return max(nums)
    raise ValueError(f"Unknown aggregate op: {op!r}")


def _now_iso() -> str:
    return _dt.datetime.now(_dt.UTC).isoformat()


def execute(query: PullQuery, ctx, result) -> None:
    """The connector entry point registered with the framework."""
    cfg = query.config or {}
    drive_file_id: str = cfg["drive_file_id"]
    sheet_name: str = cfg["sheet"]

    xlsx_bytes = _download_xlsx(drive_file_id, ctx.user_drive_token)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=False)
    try:
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"Sheet {sheet_name!r} not found in {drive_file_id}")
        ws = wb[sheet_name]

        # Resolve target → tab + (cell|None)
        from backend.app.services.connectors import parse_target
        target_tab, target_cell = parse_target(query.target)

        if target_cell is None:
            # ── Tab mode: overlay every non-empty cell ──────────────────────
            written = 0
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        result.write_cell(target_tab, cell.coordinate, cell.value)
                        written += 1
            result.provenance.append({
                "target": query.target,
                "source_kind": "xlsx_link",
                "source_ref": f"{drive_file_id}!{sheet_name} (tab)",
                "fetched_at": _now_iso(),
                "cells_written": written,
            })
            logger.info(
                "xlsx_link tab overlay: %s/%s → %s (%d cells)",
                drive_file_id, sheet_name, target_tab, written,
            )
            return

        # ── Cell or Range mode ──────────────────────────────────────────────
        if "range" in cfg:
            agg_op = cfg.get("aggregate", "sum")
            cells = ws[cfg["range"]]
            # cells is a tuple of tuples (rows of cells)
            values: list[Any] = []
            for row in cells:
                # row is a tuple of cells (when range is multi-cell) or a single cell
                if isinstance(row, tuple):
                    values.extend(c.value for c in row)
                else:
                    values.append(row.value)
            value = _aggregate(values, agg_op)
            source_ref = f"{drive_file_id}!{sheet_name}!{cfg['range']} ({agg_op})"
        else:
            cell_ref = cfg["cell"]
            value = ws[cell_ref].value
            source_ref = f"{drive_file_id}!{sheet_name}!{cell_ref}"

        if value is None and query.fallback is not None:
            value = query.fallback
            result.warnings.append(
                f"xlsx_link {source_ref} returned None; using fallback {query.fallback!r}"
            )
        result.write_cell(target_tab, target_cell, value)
        result.provenance.append({
            "target": query.target,
            "source_kind": "xlsx_link",
            "source_ref": source_ref,
            "fetched_at": _now_iso(),
            "value": value,
        })
        logger.info("xlsx_link cell: %s = %r", source_ref, value)
    finally:
        wb.close()


# Register at import time. The connectors package's __init__ imports this
# module, which triggers the registration call below as a side effect.
from backend.app.services.connectors import register as _register  # noqa: E402

_register("xlsx_link", execute)
