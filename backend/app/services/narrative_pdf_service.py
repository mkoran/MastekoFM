"""Sprint D-2 — Narrative PDF rendering from a Google Doc template.

Pipeline (all per Run):

    Google Doc template (Drive)
        │  Drive.files.export(mimeType=docx)            ← drive.file scope is enough
        ▼
    .docx bytes (Jinja2 placeholders intact)
        │  docxtpl + render_context()
        ▼
    .docx bytes (placeholders replaced with Model output values)
        │  excel_engine._find_libreoffice() + soffice --convert-to pdf
        ▼
    .pdf bytes
        │  drive_service.upload_file(...)
        ▼
    Per-run Drive folder

The user-facing experience: design a beautiful narrative report in
Google Docs (true WYSIWYG), drop in placeholders like
``{{ model.O_Results.B2 }}`` or ``{{ run.project_name }}``, save the
Doc's file id on the OutputTemplate, and every Run produces a polished
PDF with live values.

We deliberately avoid the Google Docs API (would require a new OAuth
scope: ``documents``). Round-tripping through .docx + docxtpl gives us
the same outcome with what we already have.
"""
from __future__ import annotations

import io
import logging
import os
import subprocess
import tempfile
from typing import Any

import openpyxl
from docxtpl import DocxTemplate  # type: ignore[import-untyped]

from backend.app.services import drive_service
from backend.app.services.excel_engine import LIBREOFFICE_TIMEOUT, _find_libreoffice

logger = logging.getLogger(__name__)

DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
GOOGLE_DOC_MIME = "application/vnd.google-apps.document"


def export_google_doc_as_docx(file_id: str, user_access_token: str | None = None) -> bytes | None:
    """Export a Google Doc as .docx bytes via Drive's export endpoint.

    Returns None if the export fails (caller should treat as best-effort).
    """
    try:
        # ``download_file`` uses ``files.get_media`` which doesn't work for
        # Google-native docs. Use the export endpoint instead.
        service = drive_service._get_drive_service(user_access_token)  # noqa: SLF001
        content = service.files().export_media(  # type: ignore[union-attr]
            fileId=file_id, mimeType=DOCX_MIME,
        ).execute()
        return content
    except Exception:  # noqa: BLE001
        logger.exception("Failed to export Google Doc %s as docx", file_id)
        return None


def extract_output_values_from_xlsx(xlsx_bytes: bytes) -> dict[str, dict[str, Any]]:
    """Read every cell of every ``O_*`` and ``I_*`` tab into a nested dict.

    Returned shape::

        {
            "O_Results": {"B2": 12, "B3": 35, "B4": 47, "A1": "Sum", ...},
            "I_Numbers": {"B2": 5, "B3": 7, "A1": "a", "A2": "b", ...},
        }

    Both input and output tabs are exposed so a narrative template can
    reference either (e.g. ``"The sum of {{model.I_Numbers.B2}} and ..."``).
    Calc tabs are intentionally excluded to keep the namespace tight.
    """
    out: dict[str, dict[str, Any]] = {}
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=False)
    try:
        for ws_name in wb.sheetnames:
            if not (ws_name.startswith("O_") or ws_name.startswith("I_")):
                continue
            ws = wb[ws_name]
            tab_values: dict[str, Any] = {}
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        tab_values[cell.coordinate] = cell.value
            out[ws_name] = tab_values
    finally:
        wb.close()
    return out


def _docx_to_pdf(docx_bytes: bytes) -> bytes | None:
    """Convert a .docx to PDF via LibreOffice headless.

    Mirrors :func:`excel_engine.xlsx_to_pdf` but with the Writer module.
    Best-effort: returns None if LibreOffice is missing or conversion fails.
    """
    lo_path = _find_libreoffice()
    if not lo_path:
        logger.warning("LibreOffice not found — skipping docx→pdf conversion")
        return None

    with tempfile.TemporaryDirectory() as tmpdir:
        input_path = os.path.join(tmpdir, "input.docx")
        with open(input_path, "wb") as f:
            f.write(docx_bytes)

        try:
            env = {**os.environ, "HOME": tmpdir}
            r = subprocess.run(
                [lo_path, "--headless", "--norestore", "--writer", "--convert-to", "pdf",
                 "--outdir", tmpdir, input_path],
                capture_output=True, timeout=LIBREOFFICE_TIMEOUT, cwd=tmpdir, env=env,
            )
            logger.info("docx→pdf: exit=%d", r.returncode)
            pdf_path = os.path.join(tmpdir, "input.pdf")
            if not os.path.exists(pdf_path):
                for fname in os.listdir(tmpdir):
                    if fname.endswith(".pdf"):
                        pdf_path = os.path.join(tmpdir, fname)
                        break
                else:
                    logger.warning(
                        "docx→pdf produced no output. stderr=%s",
                        (r.stderr or b"").decode("utf-8", errors="replace")[:500],
                    )
                    return None
            with open(pdf_path, "rb") as f:
                return f.read()
        except subprocess.TimeoutExpired:
            logger.warning("docx→pdf timed out after %ds", LIBREOFFICE_TIMEOUT)
            return None


def render_narrative_pdf_from_google_doc(
    *,
    template_doc_id: str,
    output_xlsx_bytes: bytes,
    run_meta: dict[str, Any],
    user_access_token: str | None = None,
) -> bytes | None:
    """End-to-end: pull template Doc → fill with run output values → return PDF bytes.

    Returns None if any step fails — callers should treat as best-effort
    (the run still completes and the xlsx + xlsx-rendered PDF survive).

    The render context exposed to the Doc template is::

        {
            "model": {
                "O_<Tab>": {"<Cell>": <value>, ...},
                "I_<Tab>": {"<Cell>": <value>, ...},
            },
            "run": {
                "id": "<run_id>",
                "project_name": "<...>",
                "model_name": "<...>",
                "pack_name": "<...>",
                "started_at": "<ISO 8601>",
                "duration_ms": <int>,
            },
        }
    """
    docx_bytes = export_google_doc_as_docx(template_doc_id, user_access_token=user_access_token)
    if not docx_bytes:
        logger.warning("Narrative PDF: could not export template doc %s", template_doc_id)
        return None

    model_values = extract_output_values_from_xlsx(output_xlsx_bytes)
    context: dict[str, Any] = {"model": model_values, "run": run_meta}

    # Render Jinja in the .docx
    try:
        doc = DocxTemplate(io.BytesIO(docx_bytes))
        doc.render(context)
        rendered = io.BytesIO()
        doc.save(rendered)
        rendered_bytes = rendered.getvalue()
    except Exception:  # noqa: BLE001
        logger.exception("Narrative PDF: docxtpl render failed")
        return None

    return _docx_to_pdf(rendered_bytes)
