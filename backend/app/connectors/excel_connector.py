"""Excel file connector — parse .xlsx files and discover fields."""
import io
from typing import Any

from openpyxl import load_workbook

from backend.app.models.datasource import DiscoveredField, FieldMapping
from backend.app.services.datasource_sync import infer_field_type


def get_sheet_names(file_content: bytes) -> list[str]:
    """Get list of sheet names from an Excel file."""
    wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def parse_excel(file_content: bytes, sheet_name: str | None = None) -> list[dict[str, Any]]:
    """Parse Excel sheet into list of row dicts. First row = headers."""
    wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    if ws is None:
        wb.close()
        return []

    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(next(rows_iter))]
    except StopIteration:
        wb.close()
        return []

    result = []
    for row in rows_iter:
        result.append(dict(zip(headers, row, strict=False)))

    wb.close()
    return result


def discover_fields(file_content: bytes, sheet_name: str | None = None) -> list[DiscoveredField]:
    """Discover fields from Excel: headers, inferred types, sample values."""
    rows = parse_excel(file_content, sheet_name)
    if not rows:
        return []

    headers = list(rows[0].keys())
    fields = []
    for header in headers:
        values = [row.get(header) for row in rows[:20]]
        fields.append(DiscoveredField(
            name=header,
            inferred_type=infer_field_type(values),
            sample_value=values[0] if values else None,
        ))
    return fields


def fetch_data(
    file_content: bytes, mappings: list[FieldMapping], sheet_name: str | None = None
) -> dict[str, Any]:
    """Extract mapped values from Excel (uses first data row)."""
    rows = parse_excel(file_content, sheet_name)
    if not rows:
        return {}

    row = rows[0]
    result: dict[str, Any] = {}
    for mapping in mappings:
        if mapping.source_field in row:
            result[mapping.source_field] = row[mapping.source_field]
    return result
