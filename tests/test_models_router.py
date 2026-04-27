"""Tests for Excel Templates router (upload, list, get, replace, delete)."""
from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl
from fastapi.testclient import TestClient

from backend.app.main import app

client = TestClient(app)
AUTH = {"Authorization": "Bearer dev-test@example.com"}
PATCH_DB = "backend.app.routers.models.get_firestore_client"
PATCH_STORAGE_UPLOAD = "backend.app.routers.models.storage_service.upload_xlsx"
PATCH_STORAGE_DELETE = "backend.app.routers.models.storage_service.delete_blob"
PATCH_STORAGE_DOWNLOAD = "backend.app.routers.models.storage_service.download_xlsx"

FIXTURE = Path(__file__).parent / "fixtures" / "campus_adele.xlsx"


def _tiny_template_bytes() -> bytes:
    """Minimal valid template for router tests that don't need the full fixture."""
    wb = openpyxl.Workbook()
    del wb[wb.sheetnames[0]]
    wb.create_sheet("I_Inputs")
    wb["I_Inputs"]["A1"] = "val"
    wb.create_sheet("O_Outputs")
    wb.create_sheet("Calc")
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@patch("backend.app.routers.models.drive_service.upload_file", return_value="drive-file-1")
@patch("backend.app.routers.models.drive_service.ensure_model_folder", return_value="model-folder-id")
@patch("backend.app.routers.models.drive_service.ensure_workspace_folders",
       return_value={"workspace": "ws-fold", "models": "models-fold",
                     "output_templates": "tpl-fold", "projects": "proj-fold"})
@patch(PATCH_DB)
def test_upload_excel_template_classifies_tabs(mock_db, mock_ws, mock_model_folder, mock_upload):
    """Sprint G1: upload now goes to Drive in {ws}/Models/{code}/{code}_v001.xlsx."""
    mock_client = MagicMock()
    mock_db.return_value = mock_client

    # workspace lookup (user has one workspace named "personal")
    ws_doc = MagicMock()
    ws_doc.id = "ws-1"
    ws_doc.to_dict.return_value = {"code_name": "personal", "members": ["dev-test@example.com"]}

    new_doc_ref = MagicMock()
    new_doc_ref.id = "tpl-1"
    settings_snap = MagicMock()
    settings_snap.exists = True
    settings_snap.to_dict.return_value = {"drive_root_folder_id": "ROOT123"}

    def collection_side(name):
        col = MagicMock()
        if name.endswith("workspaces"):
            col.where.return_value.limit.return_value.stream.return_value = [ws_doc]
        elif name.endswith("models"):
            col.document.return_value = new_doc_ref
        elif name.endswith("settings"):
            col.document.return_value.get.return_value = settings_snap
        return col

    mock_client.collection.side_effect = collection_side

    file_bytes = _tiny_template_bytes()
    resp = client.post(
        "/api/models",
        headers={**AUTH, "X-MFM-Drive-Token": "tok"},
        files={"file": ("toy.xlsx", file_bytes,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"name": "Toy", "code_name": "toy"},
    )
    assert resp.status_code == 201, resp.text
    body = resp.json()
    assert body["input_tabs"] == ["I_Inputs"]
    assert body["output_tabs"] == ["O_Outputs"]
    assert body["calc_tabs"] == ["Calc"]
    assert body["version"] == 1
    assert body["workspace_id"] == "ws-1"
    # Drive upload was called with the canonical versioned filename
    args, kwargs = mock_upload.call_args
    assert args[1] == "toy_v001.xlsx", f"expected versioned filename, got {args[1]}"
    assert mock_upload.called


@patch(PATCH_DB)
def test_upload_rejects_file_without_input_tab(mock_db):
    mock_client = MagicMock()
    mock_doc_ref = MagicMock()
    mock_doc_ref.id = "tpl-x"
    mock_client.collection.return_value.document.return_value = mock_doc_ref
    mock_db.return_value = mock_client

    wb = openpyxl.Workbook()
    del wb[wb.sheetnames[0]]
    wb.create_sheet("Calc")
    buf = BytesIO()
    wb.save(buf)

    resp = client.post(
        "/api/models",
        headers=AUTH,
        files={"file": ("bad.xlsx", buf.getvalue(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"name": "Bad"},
    )
    assert resp.status_code == 400
    assert "I_" in resp.json()["detail"]


@patch(PATCH_DB)
def test_list_excel_templates(mock_db):
    mock_client = MagicMock()
    mock_db.return_value = mock_client
    now = datetime.now(UTC)
    mock_doc = MagicMock()
    mock_doc.id = "tpl-1"
    mock_doc.to_dict.return_value = {
        "name": "Campus Adele", "code_name": "campus_adele", "version": 1,
        "input_tabs": ["I_Inputs"], "output_tabs": ["O_Out"], "calc_tabs": ["C"],
        "created_at": now, "updated_at": now,
    }
    mock_client.collection.return_value.stream.return_value = [mock_doc]

    resp = client.get("/api/models", headers=AUTH)
    assert resp.status_code == 200
    data = resp.json()
    assert len(data) == 1
    assert data[0]["input_tab_count"] == 1


def test_upload_requires_auth():
    resp = client.post(
        "/api/models",
        files={"file": ("x.xlsx", b"fake",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"name": "x"},
    )
    assert resp.status_code == 401
