"""Tests for the M_ prefix engine extensions added in Sprint A.

Covers classify_tabs M_ bucket, validate_template per-role rules,
extract_model_outputs, and overlay_outputs_onto_template.
"""
import io
from pathlib import Path

import openpyxl
import pytest

from backend.app.services import excel_template_engine as engine

FIXTURES = Path(__file__).parent / "fixtures"
HW_MODEL = FIXTURES / "helloworld_model.xlsx"
HW_INPUTS = FIXTURES / "helloworld_inputs.xlsx"
HW_REPORT = FIXTURES / "helloworld_report.xlsx"


def test_classify_recognizes_m_prefix():
    wb = openpyxl.load_workbook(HW_REPORT)
    classes = engine.classify_tabs(wb)
    wb.close()
    assert classes["m_tabs"] == ["M_Results"]
    assert classes["output_tabs"] == ["O_Report"]
    assert classes["input_tabs"] == []
    assert classes["calc_tabs"] == []


def test_classify_case_sensitivity_includes_m():
    wb = openpyxl.Workbook()
    del wb[wb.sheetnames[0]]
    wb.create_sheet("M_Real")
    wb.create_sheet("m_fake")  # lowercase, should be calc
    wb.create_sheet("MM_Foo")  # not exactly M_, should be calc
    buf = io.BytesIO()
    wb.save(buf)
    classes = engine.classify_bytes(buf.getvalue())
    assert classes["m_tabs"] == ["M_Real"]
    assert "m_fake" in classes["calc_tabs"]
    assert "MM_Foo" in classes["calc_tabs"]


def test_validate_template_role_model_passes_on_helloworld():
    wb = openpyxl.load_workbook(HW_MODEL)
    try:
        errors = engine.validate_template(wb, role=engine.ROLE_MODEL)
    finally:
        wb.close()
    assert errors == []


def test_validate_template_role_model_rejects_pack_shape():
    wb = openpyxl.load_workbook(HW_INPUTS)
    try:
        errors = engine.validate_template(wb, role=engine.ROLE_MODEL)
    finally:
        wb.close()
    # Pack has only I_ tabs — Model needs O_ too
    assert any("O_ tabs" in e for e in errors)


def test_validate_template_role_pack_passes_on_helloworld_inputs():
    wb = openpyxl.load_workbook(HW_INPUTS)
    try:
        errors = engine.validate_template(wb, role=engine.ROLE_ASSUMPTION_PACK)
    finally:
        wb.close()
    assert errors == []


def test_validate_template_role_pack_rejects_o_tab():
    wb = openpyxl.Workbook()
    del wb[wb.sheetnames[0]]
    wb.create_sheet("I_Inputs")
    wb.create_sheet("O_Outputs")  # not allowed on a Pack
    errors = engine.validate_template(wb, role=engine.ROLE_ASSUMPTION_PACK)
    wb.close()
    assert any("not I_ tabs" in e for e in errors)


def test_validate_template_role_output_template_passes_on_helloworld_report():
    wb = openpyxl.load_workbook(HW_REPORT)
    try:
        errors = engine.validate_template(wb, role=engine.ROLE_OUTPUT_TEMPLATE_XLSX)
    finally:
        wb.close()
    assert errors == []


def test_validate_template_role_output_template_rejects_i_tab():
    wb = openpyxl.Workbook()
    del wb[wb.sheetnames[0]]
    wb.create_sheet("I_Bad")  # not allowed on an OutputTemplate
    wb.create_sheet("O_Report")
    errors = engine.validate_template(wb, role=engine.ROLE_OUTPUT_TEMPLATE_XLSX)
    wb.close()
    assert any("I_ tabs" in e for e in errors)


def test_validate_template_unknown_role():
    wb = openpyxl.Workbook()
    errors = engine.validate_template(wb, role="bogus")
    assert any("Unknown role" in e for e in errors)


def test_basename_strips_known_prefixes():
    assert engine.basename("I_Foo") == "Foo"
    assert engine.basename("O_Bar") == "Bar"
    assert engine.basename("M_Baz") == "Baz"
    assert engine.basename("PlainTab") == "PlainTab"


def test_extract_model_outputs_reads_o_tabs(template_recalced_model):
    outputs = engine.extract_model_outputs(template_recalced_model)
    assert "O_Results" in outputs
    # The recalced model has sum=2+3=5, product=2*3=6 (default literals from helloworld_model)
    assert outputs["O_Results"]["A1"] == "sum"
    assert outputs["O_Results"]["A2"] == "product"
    assert outputs["O_Results"]["B1"] == 5
    assert outputs["O_Results"]["B2"] == 6


def test_overlay_outputs_writes_into_m_tabs():
    tpl_bytes = HW_REPORT.read_bytes()
    model_outputs = {
        "O_Results": {"A1": "sum", "B1": 12, "A2": "product", "B2": 35},
    }
    merged, warnings = engine.overlay_outputs_onto_template(tpl_bytes, model_outputs)
    assert warnings == []

    wb = openpyxl.load_workbook(io.BytesIO(merged))
    try:
        assert wb["M_Results"]["B1"].value == 12
        assert wb["M_Results"]["B2"].value == 35
    finally:
        wb.close()


def test_overlay_outputs_warns_on_missing_match():
    tpl_bytes = HW_REPORT.read_bytes()
    model_outputs = {
        "O_Other": {"A1": "x", "B1": 1},  # no matching M_Other in template
    }
    _, warnings = engine.overlay_outputs_onto_template(tpl_bytes, model_outputs)
    # Template has M_Results but model_outputs has no O_Results -> should warn
    assert any("M_Results" in w for w in warnings)


@pytest.fixture
def template_recalced_model():
    """Recalculate the Hello World Model via LibreOffice and return its bytes.

    Skipped if LibreOffice isn't installed in the test environment.
    """
    from backend.app.services import excel_engine

    if not excel_engine._find_libreoffice():
        pytest.skip("LibreOffice not available in this test environment")

    bytes_in = HW_MODEL.read_bytes()
    recalced = excel_engine.recalculate_with_libreoffice(bytes_in)
    if recalced is None:
        pytest.skip("LibreOffice failed to recalculate Hello World")
    return recalced
