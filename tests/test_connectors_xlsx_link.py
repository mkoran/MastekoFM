"""Sprint I-1 — XLSX-Link connector + framework unit tests.

Mocks the Drive download so the connector code path is exercised without
network access. Confirms cell mode, range+aggregate mode, and tab-overlay
mode all produce the expected ``cell_writes`` and ``provenance``.
"""
from __future__ import annotations

import io
from unittest.mock import patch

import openpyxl

from backend.app.models.assumption_pack import PullQuery, PullSpec
from backend.app.services import connectors
from backend.app.services.pack_synth import synthesize_pack_xlsx_from_overrides


def _make_source_bytes(sheet_cells: dict[str, dict[str, object]]) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, cells in sheet_cells.items():
        ws = wb.create_sheet(name)
        for ref, val in cells.items():
            ws[ref] = val
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _ctx() -> connectors.ConnectorContext:
    return connectors.ConnectorContext(
        user_drive_token="fake-token", workspace_id="ws-1", run_id="r-1",
    )


def test_xlsx_link_cell_mode():
    src = _make_source_bytes({"Master": {"B1": 5, "B2": 7}})
    spec = PullSpec(queries=[
        PullQuery(
            target="I_Numbers.B1", kind="xlsx_link",
            config={"drive_file_id": "abc", "sheet": "Master", "cell": "B1"},
        ),
    ])
    with patch("backend.app.services.drive_service.download_file", return_value=src):
        result = connectors.execute_pull_spec(spec, _ctx())
    assert result.cell_writes == {"I_Numbers": {"B1": 5}}
    assert len(result.provenance) == 1
    assert result.provenance[0]["target"] == "I_Numbers.B1"


def test_xlsx_link_range_mode_with_sum():
    src = _make_source_bytes({"Sheet1": {"B1": 10, "B2": 20, "B3": 30}})
    spec = PullSpec(queries=[
        PullQuery(
            target="I_Total.B1", kind="xlsx_link",
            config={"drive_file_id": "abc", "sheet": "Sheet1",
                    "range": "B1:B3", "aggregate": "sum"},
        ),
    ])
    with patch("backend.app.services.drive_service.download_file", return_value=src):
        result = connectors.execute_pull_spec(spec, _ctx())
    assert result.cell_writes == {"I_Total": {"B1": 60}}


def test_xlsx_link_tab_mode_overlays_every_cell():
    src = _make_source_bytes({"Master": {"A1": "a", "B1": 5, "A2": "b", "B2": 7}})
    spec = PullSpec(queries=[
        PullQuery(
            target="I_Numbers", kind="xlsx_link",
            config={"drive_file_id": "abc", "sheet": "Master"},
        ),
    ])
    with patch("backend.app.services.drive_service.download_file", return_value=src):
        result = connectors.execute_pull_spec(spec, _ctx())
    assert result.cell_writes["I_Numbers"] == {"A1": "a", "B1": 5, "A2": "b", "B2": 7}
    assert result.provenance[0]["cells_written"] == 4


def test_synthesize_round_trip_through_engine_compatible_xlsx():
    """Confirm the synthesized xlsx loads back with the same I_* cell values."""
    overrides = {
        "I_Numbers": {"B1": 5, "B2": 7},
        "I_Pricing": {"B5": 1500.0, "B6": "Quarterly"},
    }
    pack_bytes = synthesize_pack_xlsx_from_overrides(overrides)
    wb = openpyxl.load_workbook(io.BytesIO(pack_bytes))
    assert set(wb.sheetnames) == {"I_Numbers", "I_Pricing"}
    assert wb["I_Numbers"]["B1"].value == 5
    assert wb["I_Numbers"]["B2"].value == 7
    assert wb["I_Pricing"]["B5"].value == 1500.0
    assert wb["I_Pricing"]["B6"].value == "Quarterly"


def test_pull_spec_warn_on_error_uses_fallback():
    """When the connector raises and on_error='warn', fallback is written."""
    spec = PullSpec(
        on_error="warn",
        queries=[
            PullQuery(
                target="I_Numbers.B1", kind="xlsx_link",
                config={"drive_file_id": "abc", "sheet": "Master", "cell": "B1"},
                fallback=999,
            ),
        ],
    )
    # Drive download returns None → connector raises RuntimeError → handled
    with patch("backend.app.services.drive_service.download_file", return_value=None):
        result = connectors.execute_pull_spec(spec, _ctx())
    assert result.cell_writes == {"I_Numbers": {"B1": 999}}
    assert any("xlsx_link" in w for w in result.warnings)


def test_pull_spec_unregistered_kind_uses_fallback(monkeypatch):
    """If a kind is in the model's Literal but not in the registry,
    execute_pull_spec catches the KeyError and falls back gracefully.

    Simulate an unregistered kind by temporarily emptying the registry
    for our query's kind. (PullQuery.kind is constrained by Pydantic so we
    can't pass a literal string that's not in the Literal.)
    """
    spec = PullSpec(
        on_error="warn",
        queries=[
            PullQuery(
                target="I_Foo.B1", kind="airtable",
                config={},
                fallback=42,
            ),
        ],
    )
    # Temporarily de-register airtable so the framework can't resolve it.
    saved = connectors._REGISTRY.pop("airtable", None)
    try:
        result = connectors.execute_pull_spec(spec, _ctx())
    finally:
        if saved is not None:
            connectors._REGISTRY["airtable"] = saved
    assert result.cell_writes == {"I_Foo": {"B1": 42}}
    assert any("unknown connector" in w.lower() for w in result.warnings)


def test_parse_target_cell_vs_tab():
    assert connectors.parse_target("I_Numbers.B1") == ("I_Numbers", "B1")
    assert connectors.parse_target("I_Numbers") == ("I_Numbers", None)
