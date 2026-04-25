"""Sprint UX-01 regression tests for AssumptionPacks router.

Covers two production bugs surfaced post-Sprint-B:
  Bug UX-01-01: POST /api/projects/{id}/assumption-packs returned 500 because the
                handler still read the legacy `template_id` field (Sprint B
                renamed it to `default_model_id`).
  Bug UX-01-02: POST .../{pack_id}/calculate failed for the same reason; this
                test asserts the happy path now succeeds end-to-end with the
                new field.

These tests are the regression gate: a future revert of the field rename would
break them on CI before any deploy.
"""
from io import BytesIO
from unittest.mock import MagicMock, patch

import openpyxl
from fastapi.testclient import TestClient

from backend.app.main import app

client = TestClient(app)
AUTH = {"Authorization": "Bearer dev-test@example.com"}
PATCH_DB = "backend.app.routers.assumption_packs.get_firestore_client"
PATCH_LOAD_MODEL = "backend.app.routers.assumption_packs.pack_store.load_model_bytes_compat"
PATCH_GET_STORE = "backend.app.routers.assumption_packs.pack_store.get_store"
PATCH_EXTRACT = (
    "backend.app.routers.assumption_packs.excel_template_engine.extract_scenario_from_template"
)
PATCH_CALCULATE = "backend.app.routers.assumption_packs.excel_template_engine.calculate"
PATCH_UPLOAD_XLSX = "backend.app.routers.assumption_packs.storage_service.upload_xlsx"


def _tiny_xlsx_bytes() -> bytes:
    wb = openpyxl.Workbook()
    del wb[wb.sheetnames[0]]
    wb.create_sheet("I_Numbers")
    wb["I_Numbers"]["A1"] = 5
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Helpers ──────────────────────────────────────────────────────────────────


def _project_with_default_model(default_model_id: str = "model-1") -> dict:
    return {
        "name": "Hello World",
        "code_name": "helloworld",
        "default_model_id": default_model_id,
        "default_model_name": "Hello World Model",
        "default_model_version": 1,
        "status": "active",
    }


def _gcs_model() -> dict:
    return {
        "name": "Hello World Model",
        "code_name": "helloworld_model",
        "version": 1,
        "storage_path": "models/m1/v1.xlsx",
        "drive_file_id": None,
    }


def _drive_model() -> dict:
    return {
        "name": "Hello World Model",
        "code_name": "helloworld_model",
        "version": 1,
        "storage_path": None,
        "drive_file_id": "drive-file-abc",
    }


# ── UX-01-01: Create AssumptionPack regression ───────────────────────────────


@patch(PATCH_DB)
def test_create_pack_returns_400_when_project_has_no_default_model(mock_db):
    """Bug UX-01-01 prequel: a Project with no default Model must return 400 (not 500).

    Pre-fix this hit a Firestore lookup with empty string and crashed.
    """
    mock_client = MagicMock()
    mock_db.return_value = mock_client

    proj_doc = MagicMock()
    proj_doc.exists = True
    proj_doc.to_dict.return_value = {"name": "Empty Project", "code_name": "ep"}
    mock_client.collection.return_value.document.return_value.get.return_value = proj_doc

    resp = client.post(
        "/api/projects/proj-1/assumption-packs",
        json={"name": "s1"},
        headers=AUTH,
    )
    assert resp.status_code == 400, resp.text
    assert "default Model" in resp.json()["detail"]


@patch(PATCH_DB)
def test_create_pack_legacy_template_id_field_still_works(mock_db):
    """Belt-and-suspenders: an un-migrated Project with only `template_id` still works."""
    mock_client = MagicMock()
    mock_db.return_value = mock_client

    proj_doc = MagicMock()
    proj_doc.exists = True
    proj_doc.to_dict.return_value = {
        "name": "Legacy",
        "code_name": "legacy",
        "template_id": "model-legacy",  # old field, no default_model_id
    }
    model_doc = MagicMock()
    model_doc.exists = True
    model_doc.to_dict.return_value = _gcs_model()

    settings_doc = MagicMock()
    settings_doc.exists = True
    settings_doc.to_dict.return_value = {"default_scenario_storage_kind": "gcs"}

    new_pack_ref = MagicMock()
    new_pack_ref.id = "pack-legacy"

    def collection_side_effect(name):
        col = MagicMock()
        if name.endswith("projects"):
            doc_obj = MagicMock()
            doc_obj.get.return_value = proj_doc
            sub = MagicMock()
            sub.document.return_value = new_pack_ref
            doc_obj.collection.return_value = sub
            col.document.return_value = doc_obj
        elif name.endswith("models"):
            doc_obj = MagicMock()
            doc_obj.get.return_value = model_doc
            col.document.return_value = doc_obj
        elif name.endswith("settings"):
            doc_obj = MagicMock()
            doc_obj.get.return_value = settings_doc
            col.document.return_value = doc_obj
        return col

    mock_client.collection.side_effect = collection_side_effect

    with patch(PATCH_LOAD_MODEL, return_value=_tiny_xlsx_bytes()), \
         patch(PATCH_EXTRACT, return_value=_tiny_xlsx_bytes()), \
         patch(PATCH_GET_STORE) as mock_get_store, \
         patch(PATCH_UPLOAD_XLSX, return_value="https://example/x.xlsx"):
        store = MagicMock()
        store.kind = "gcs"
        store.write_bytes.return_value = {
            "storage_kind": "gcs",
            "storage_path": "p",
            "drive_file_id": None,
            "size_bytes": 100,
        }
        store.open_url.return_value = "https://example/x.xlsx"
        mock_get_store.return_value = store

        resp = client.post(
            "/api/projects/proj-legacy/assumption-packs",
            json={"name": "s1", "storage_kind": "gcs"},
            headers=AUTH,
        )
    assert resp.status_code == 201, resp.text
    assert resp.json()["name"] == "s1"


@patch(PATCH_DB)
def test_create_pack_post_sprint_b_default_model_id_works(mock_db):
    """UX-01-01 fix: the canonical post-Sprint-B field path returns 201."""
    mock_client = MagicMock()
    mock_db.return_value = mock_client

    proj_doc = MagicMock()
    proj_doc.exists = True
    proj_doc.to_dict.return_value = _project_with_default_model("m1")
    model_doc = MagicMock()
    model_doc.exists = True
    model_doc.to_dict.return_value = _gcs_model()

    settings_doc = MagicMock()
    settings_doc.exists = True
    settings_doc.to_dict.return_value = {"default_scenario_storage_kind": "gcs"}

    new_pack_ref = MagicMock()
    new_pack_ref.id = "pack-new"

    def collection_side_effect(name):
        col = MagicMock()
        if name.endswith("projects"):
            doc_obj = MagicMock()
            doc_obj.get.return_value = proj_doc
            sub = MagicMock()
            sub.document.return_value = new_pack_ref
            doc_obj.collection.return_value = sub
            col.document.return_value = doc_obj
        elif name.endswith("models"):
            doc_obj = MagicMock()
            doc_obj.get.return_value = model_doc
            col.document.return_value = doc_obj
        elif name.endswith("settings"):
            doc_obj = MagicMock()
            doc_obj.get.return_value = settings_doc
            col.document.return_value = doc_obj
        return col

    mock_client.collection.side_effect = collection_side_effect

    with patch(PATCH_LOAD_MODEL, return_value=_tiny_xlsx_bytes()), \
         patch(PATCH_EXTRACT, return_value=_tiny_xlsx_bytes()), \
         patch(PATCH_GET_STORE) as mock_get_store, \
         patch(PATCH_UPLOAD_XLSX, return_value="https://example/x.xlsx"):
        store = MagicMock()
        store.kind = "gcs"
        store.write_bytes.return_value = {
            "storage_kind": "gcs",
            "storage_path": "p",
            "drive_file_id": None,
            "size_bytes": 100,
        }
        store.open_url.return_value = "https://example/x.xlsx"
        mock_get_store.return_value = store

        resp = client.post(
            "/api/projects/proj-1/assumption-packs",
            json={"name": "s1", "storage_kind": "gcs"},
            headers=AUTH,
        )
    assert resp.status_code == 201, resp.text


# ── UX-01-02: Calculate regression ───────────────────────────────────────────


@patch(PATCH_DB)
def test_calculate_returns_400_when_project_has_no_default_model(mock_db):
    """Calculate must surface 400, not 500, when no default Model is bound."""
    mock_client = MagicMock()
    mock_db.return_value = mock_client

    proj_doc = MagicMock()
    proj_doc.exists = True
    proj_doc.to_dict.return_value = {"name": "Empty", "code_name": "e"}

    mock_client.collection.return_value.document.return_value.get.return_value = proj_doc

    resp = client.post(
        "/api/projects/proj-1/assumption-packs/scn-1/calculate",
        json={},
        headers=AUTH,
    )
    assert resp.status_code == 400


@patch(PATCH_DB)
def test_calculate_happy_path_with_default_model_id(mock_db):
    """UX-01-02 fix: Calculate runs end-to-end against post-Sprint-B Project shape."""
    mock_client = MagicMock()
    mock_db.return_value = mock_client

    proj_doc = MagicMock()
    proj_doc.exists = True
    proj_doc.to_dict.return_value = _project_with_default_model("m1")
    model_doc = MagicMock()
    model_doc.exists = True
    model_doc.to_dict.return_value = _gcs_model()

    pack_data = {
        "name": "Hello World Inputs",
        "code_name": "helloworld_inputs",
        "version": 1,
        "storage_kind": "gcs",
        "storage_path": "packs/p1/v1.xlsx",
        "drive_file_id": None,
    }
    pack_doc = MagicMock()
    pack_doc.exists = True
    pack_doc.to_dict.return_value = pack_data

    new_run_ref = MagicMock()
    new_run_ref.id = "run-1"

    def collection_side_effect(name):
        col = MagicMock()
        if name.endswith("projects"):
            proj_doc_obj = MagicMock()
            proj_doc_obj.get.return_value = proj_doc
            sub = MagicMock()
            sub_doc = MagicMock()
            sub_doc.get.return_value = pack_doc
            run_sub = MagicMock()
            run_sub.document.return_value = new_run_ref
            sub_doc.collection.return_value = run_sub
            sub.document.return_value = sub_doc
            proj_doc_obj.collection.return_value = sub
            col.document.return_value = proj_doc_obj
        elif name.endswith("models"):
            doc_obj = MagicMock()
            doc_obj.get.return_value = model_doc
            col.document.return_value = doc_obj
        return col

    mock_client.collection.side_effect = collection_side_effect

    with patch(PATCH_LOAD_MODEL, return_value=_tiny_xlsx_bytes()), \
         patch(PATCH_GET_STORE) as mock_get_store, \
         patch(PATCH_CALCULATE) as mock_calc, \
         patch(PATCH_UPLOAD_XLSX, return_value="https://example/out.xlsx"):
        store = MagicMock()
        store.kind = "gcs"
        store.read_bytes.return_value = _tiny_xlsx_bytes()
        store.open_url.return_value = "https://example/in.xlsx"
        mock_get_store.return_value = store

        mock_calc.return_value = {
            "output_bytes": _tiny_xlsx_bytes(),
            "merged_bytes": _tiny_xlsx_bytes(),
            "warnings": [],
            "recalculated": True,
        }

        resp = client.post(
            "/api/projects/proj-1/assumption-packs/scn-1/calculate",
            json={},
            headers=AUTH,
        )

    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["status"] == "done"
    # template_version_used should come from the Model (was reading missing tpl before fix)
    assert body["template_version_used"] == 1
    assert mock_calc.called, "engine.calculate should have been invoked"
