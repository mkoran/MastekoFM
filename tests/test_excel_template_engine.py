"""Tests for the Excel Template engine (tab classification, overlay, calculate).

Uses the real Campus Adele model as the fixture, matching the spike that
proved the architecture works end-to-end.
"""
import io
from pathlib import Path

import openpyxl
import pytest

from backend.app.services import excel_template_engine as engine

FIXTURE = Path(__file__).parent / "fixtures" / "campus_adele.xlsx"


@pytest.fixture
def template_bytes() -> bytes:
    return FIXTURE.read_bytes()


def test_classify_tabs_campus_adele(template_bytes):
    classes = engine.classify_bytes(template_bytes)
    # Case-sensitive: "i_Cap Table" is NOT an input
    assert "i_Cap Table" in classes["calc_tabs"]
    assert "I_Inputs & Assumptions" in classes["input_tabs"]
    assert "I_Budget Input Data" in classes["input_tabs"]
    assert "I_Unit Rent Roll Data" in classes["input_tabs"]
    assert "I_Preferred Equity" in classes["input_tabs"]
    assert "I_Mezzanine Debt" in classes["input_tabs"]
    assert "O_Annual Summary" in classes["output_tabs"]
    assert len(classes["input_tabs"]) == 5
    assert len(classes["output_tabs"]) == 1


def test_classify_case_sensitive():
    # Build a tiny workbook with mixed-case prefixes
    wb = openpyxl.Workbook()
    del wb[wb.sheetnames[0]]
    wb.create_sheet("I_Real")
    wb.create_sheet("i_Fake")
    wb.create_sheet("O_Out")
    wb.create_sheet("o_nope")
    buf = io.BytesIO()
    wb.save(buf)
    classes = engine.classify_bytes(buf.getvalue())
    assert classes["input_tabs"] == ["I_Real"]
    assert classes["output_tabs"] == ["O_Out"]
    assert set(classes["calc_tabs"]) == {"i_Fake", "o_nope"}


def test_validate_template_requires_input_tab():
    wb = openpyxl.Workbook()
    del wb[wb.sheetnames[0]]
    wb.create_sheet("Calcs")
    buf = io.BytesIO()
    wb.save(buf)
    loaded = openpyxl.load_workbook(io.BytesIO(buf.getvalue()))
    errors = engine.validate_template(loaded)
    assert any("I_ tabs" in e for e in errors)


def test_validate_template_ok(template_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    try:
        assert engine.validate_template(wb) == []
    finally:
        wb.close()


def test_extract_scenario_has_only_input_tabs(template_bytes):
    sc_bytes = engine.extract_scenario_from_template(template_bytes)
    classes = engine.classify_bytes(sc_bytes)
    assert set(classes["input_tabs"]) == {
        "I_Inputs & Assumptions",
        "I_Budget Input Data",
        "I_Unit Rent Roll Data",
        "I_Preferred Equity",
        "I_Mezzanine Debt",
    }
    assert classes["output_tabs"] == []
    assert classes["calc_tabs"] == []
    # Scenario file should be materially smaller than the template
    assert len(sc_bytes) < len(template_bytes) / 2


def test_overlay_preserves_calc_formulas(template_bytes):
    """Overlaying a scenario must not disturb calc tab formulas."""
    scenario_bytes = engine.extract_scenario_from_template(template_bytes)
    merged, warnings = engine.overlay_scenario_on_template(template_bytes, scenario_bytes)
    assert warnings == []
    # Original formula counts proven during spike
    wb = openpyxl.load_workbook(io.BytesIO(merged))
    try:
        assert sum(
            1
            for row in wb["Construction Budget & Draws"].iter_rows()
            for c in row
            if isinstance(c.value, str) and c.value.startswith("=")
        ) == 7302
        assert sum(
            1
            for row in wb["O_Annual Summary"].iter_rows()
            for c in row
            if isinstance(c.value, str) and c.value.startswith("=")
        ) == 493
    finally:
        wb.close()


def test_overlay_mutation_propagates_to_cells(template_bytes):
    """If scenario mutates an input cell, the merged file has the new value."""
    scenario_bytes = engine.extract_scenario_from_template(template_bytes)
    wb = openpyxl.load_workbook(io.BytesIO(scenario_bytes))
    try:
        wb["I_Inputs & Assumptions"]["B8"].value = 14.3  # Construction Duration
        buf = io.BytesIO()
        wb.save(buf)
        scenario_bytes = buf.getvalue()
    finally:
        wb.close()

    merged, _ = engine.overlay_scenario_on_template(template_bytes, scenario_bytes)
    wb_merged = openpyxl.load_workbook(io.BytesIO(merged))
    try:
        assert wb_merged["I_Inputs & Assumptions"]["B8"].value == 14.3
    finally:
        wb_merged.close()


def test_overlay_rejects_scenario_with_calc_tab(template_bytes):
    scenario_bytes = engine.extract_scenario_from_template(template_bytes)
    wb = openpyxl.load_workbook(io.BytesIO(scenario_bytes))
    try:
        wb.create_sheet("Illegal Calc Tab")
        buf = io.BytesIO()
        wb.save(buf)
        scenario_bytes = buf.getvalue()
    finally:
        wb.close()
    with pytest.raises(ValueError, match="not I_ tabs"):
        engine.overlay_scenario_on_template(template_bytes, scenario_bytes)


def test_overlay_rejects_scenario_missing_required_input(template_bytes):
    scenario_bytes = engine.extract_scenario_from_template(template_bytes)
    wb = openpyxl.load_workbook(io.BytesIO(scenario_bytes))
    try:
        del wb["I_Budget Input Data"]
        buf = io.BytesIO()
        wb.save(buf)
        scenario_bytes = buf.getvalue()
    finally:
        wb.close()
    with pytest.raises(ValueError, match="missing required input tabs"):
        engine.overlay_scenario_on_template(template_bytes, scenario_bytes)


def test_replace_template_tabs_reports_diff(template_bytes):
    """Replacing a template produces a per-tab diff report (added/removed/kept)."""
    # Build a "new version" that drops one I_ tab and adds another
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    try:
        del wb["I_Mezzanine Debt"]
        wb.create_sheet("I_New Source")
        buf = io.BytesIO()
        wb.save(buf)
        new_bytes = buf.getvalue()
    finally:
        wb.close()

    _, report = engine.replace_template_tabs(template_bytes, new_bytes)
    assert "I_New Source" in report["input_tabs"]["added"]
    assert "I_Mezzanine Debt" in report["input_tabs"]["removed"]
    assert "I_Inputs & Assumptions" in report["input_tabs"]["kept"]
