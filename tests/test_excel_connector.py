"""Tests for Excel connector."""
import io

from openpyxl import Workbook

from backend.app.connectors.excel_connector import discover_fields, fetch_data, get_sheet_names, parse_excel
from backend.app.models.datasource import FieldMapping


def _make_xlsx() -> bytes:
    """Create a simple test Excel file in memory."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["name", "amount", "rate"])
    ws.append(["Land Cost", 2500000, 0.05])
    ws.append(["Rent", 1850, 0.035])

    # Add second sheet
    ws2 = wb.create_sheet("Summary")
    ws2.append(["total", 100000])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_get_sheet_names():
    content = _make_xlsx()
    names = get_sheet_names(content)
    assert "Data" in names
    assert "Summary" in names


def test_parse_excel():
    content = _make_xlsx()
    rows = parse_excel(content, "Data")
    assert len(rows) == 2
    assert rows[0]["name"] == "Land Cost"
    assert rows[0]["amount"] == 2500000


def test_discover_fields():
    content = _make_xlsx()
    fields = discover_fields(content, "Data")
    assert len(fields) == 3
    names = [f.name for f in fields]
    assert "name" in names
    assert "amount" in names

    amount_field = next(f for f in fields if f.name == "amount")
    assert amount_field.inferred_type == "number"


def test_fetch_data():
    content = _make_xlsx()
    mappings = [FieldMapping(source_field="amount", assumption_key="land_cost")]
    data = fetch_data(content, mappings, "Data")
    assert data["amount"] == 2500000


def test_empty_excel():
    wb = Workbook()
    buf = io.BytesIO()
    wb.save(buf)
    fields = discover_fields(buf.getvalue())
    assert fields == []
