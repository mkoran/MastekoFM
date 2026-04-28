"""Sprint INFRA-002 unit tests for the e2e smoke script.

The script itself is run end-to-end against a live API in CI; these are
mock-only tests of its branching logic so a future refactor doesn't silently
break the skip / required / assert paths.
"""
import importlib.util
import io
import sys
from pathlib import Path
from unittest.mock import patch

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[1]
SCRIPT_PATH = REPO_ROOT / "scripts" / "smoke" / "e2e_run_smoke.py"


def _load_script_module():
    spec = importlib.util.spec_from_file_location("_e2e_smoke", SCRIPT_PATH)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _build_correct_output_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("O_Report")
    ws["A3"] = "Sum:"
    ws["B3"] = 12
    ws["A4"] = "Product:"
    ws["B4"] = 35
    ws["A5"] = "Total:"
    ws["B5"] = 47
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_wrong_output_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("O_Report")
    ws["B3"] = 99  # all wrong
    ws["B4"] = 99
    ws["B5"] = 99
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


SEED_RESPONSE = {
    "project_id": "p-1",
    "model_id": "m-1",
    "assumption_pack_id": "pack-1",
    "output_template_id": "tpl-1",
}


def _patched_request_factory(seed_resp, run_resp, get_run_resp=None):
    """Build a side-effect for module._request that responds based on URL."""
    get_run_resp = get_run_resp or run_resp

    def side_effect(method, url, **kwargs):
        if method == "POST" and url.endswith("/api/seed/helloworld"):
            return seed_resp
        if method == "POST" and url.endswith("/api/runs"):
            return run_resp
        if method == "GET" and "/api/runs/" in url:
            return get_run_resp
        raise AssertionError(f"Unexpected request: {method} {url}")

    return side_effect


def test_run_e2e_passes_when_output_is_correct():
    mod = _load_script_module()
    run_resp = {
        "id": "r-1",
        "status": "completed",
        "duration_ms": 100,
        "output_download_url": "https://example/out.xlsx",
    }
    with patch.object(mod, "_request", side_effect=_patched_request_factory(SEED_RESPONSE, run_resp)), \
         patch.object(mod, "_download", return_value=_build_correct_output_xlsx()):
        rc = mod.run_e2e("https://api", "auth", "drive")
    assert rc == 0


def test_run_e2e_fails_when_run_status_is_failed():
    mod = _load_script_module()
    run_resp = {
        "id": "r-1",
        "status": "failed",
        "error": "engine exploded",
        "output_download_url": None,
    }
    with patch.object(mod, "_request", side_effect=_patched_request_factory(SEED_RESPONSE, run_resp)):
        rc = mod.run_e2e("https://api", "auth", "drive")
    assert rc == 1


def test_run_e2e_soft_passes_when_no_output_url():
    """Sprint G1: SA storage-quota limitation means the run output may be null.
    The engine ran (status=completed), so we treat this as a soft pass."""
    mod = _load_script_module()
    run_resp = {"id": "r-1", "status": "completed", "output_download_url": None}
    with patch.object(mod, "_request", side_effect=_patched_request_factory(SEED_RESPONSE, run_resp)):
        rc = mod.run_e2e("https://api", "auth", "drive")
    assert rc == 0  # soft pass with warning


def test_run_e2e_fails_when_cells_wrong():
    mod = _load_script_module()
    run_resp = {
        "id": "r-1", "status": "completed",
        "output_download_url": "https://example/out.xlsx",
    }
    with patch.object(mod, "_request", side_effect=_patched_request_factory(SEED_RESPONSE, run_resp)), \
         patch.object(mod, "_download", return_value=_build_wrong_output_xlsx()):
        rc = mod.run_e2e("https://api", "auth", "drive")
    assert rc == 1


def test_run_e2e_polls_until_completed():
    mod = _load_script_module()
    pending = {"id": "r-1", "status": "running", "output_download_url": None}
    completed = {
        "id": "r-1", "status": "completed",
        "output_download_url": "https://example/out.xlsx",
    }
    state = {"call_count": 0}

    def request_side_effect(method, url, **kwargs):
        if method == "POST" and url.endswith("/api/seed/helloworld"):
            return SEED_RESPONSE
        if method == "POST" and url.endswith("/api/runs"):
            return pending  # initial POST returns running
        if method == "GET" and "/api/runs/" in url:
            state["call_count"] += 1
            return completed if state["call_count"] >= 2 else pending
        raise AssertionError(f"Unexpected: {method} {url}")

    with patch.object(mod, "_request", side_effect=request_side_effect), \
         patch.object(mod, "_download", return_value=_build_correct_output_xlsx()), \
         patch.object(mod.time, "sleep"):  # don't actually sleep in tests
        rc = mod.run_e2e("https://api", "auth", "drive")
    assert rc == 0
    assert state["call_count"] >= 2


def test_main_skips_when_no_token_and_not_required(monkeypatch):
    mod = _load_script_module()
    monkeypatch.setenv("API_BASE_URL", "https://api")
    monkeypatch.setenv("AUTH_TOKEN", "dev-x@y")
    monkeypatch.delenv("MFM_DRIVE_TOKEN", raising=False)
    with patch.object(mod, "get_drive_token", return_value=None), \
         patch.object(sys, "argv", ["e2e"]):
        rc = mod.main()
    assert rc == 0  # skip, not fail


def test_main_fails_when_no_token_and_required(monkeypatch):
    mod = _load_script_module()
    monkeypatch.setenv("API_BASE_URL", "https://api")
    monkeypatch.setenv("AUTH_TOKEN", "dev-x@y")
    monkeypatch.delenv("MFM_DRIVE_TOKEN", raising=False)
    with patch.object(mod, "get_drive_token", return_value=None), \
         patch.object(sys, "argv", ["e2e", "--required"]):
        rc = mod.main()
    assert rc == 2  # required but missing → fail
