"""End-to-end tests for services/run_executor.py — Hello World pipeline.

Requires LibreOffice. Tests are skipped if it's not available locally.
"""
import io
from pathlib import Path

import openpyxl
import pytest

from backend.app.services import excel_engine, run_executor

FIXTURES = Path(__file__).parent / "fixtures"
HW_MODEL = FIXTURES / "helloworld_model.xlsx"
HW_INPUTS = FIXTURES / "helloworld_inputs.xlsx"
HW_REPORT = FIXTURES / "helloworld_report.xlsx"


@pytest.fixture(autouse=True)
def _require_libreoffice():
    if not excel_engine._find_libreoffice():
        pytest.skip("LibreOffice not available in this test environment")


def test_helloworld_run_produces_expected_numbers():
    result = run_executor.execute_run_sync(
        model_bytes=HW_MODEL.read_bytes(),
        pack_bytes=HW_INPUTS.read_bytes(),
        output_template_bytes=HW_REPORT.read_bytes(),
        output_template_format="xlsx",
    )

    assert result["stage1_recalculated"] is True
    assert result["stage2_recalculated"] is True
    assert result["warnings"] == []

    # Stage 1: Model O_Results should have sum=12 (5+7), product=35 (5*7)
    assert result["model_outputs"]["O_Results"]["B1"] == 12
    assert result["model_outputs"]["O_Results"]["B2"] == 35

    # Stage 2: write the artifact and verify O_Report values
    import os
    import tempfile
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(result["output_bytes"])
        tmp_path = tmp.name
    try:
        wb = openpyxl.load_workbook(tmp_path, data_only=True)
        try:
            ws = wb["O_Report"]
            assert ws["B3"].value == 12, "Sum cell should be 12"
            assert ws["B4"].value == 35, "Product cell should be 35"
            assert ws["B5"].value == 47, "Total cell should be 47 (12+35)"
        finally:
            wb.close()
    finally:
        os.unlink(tmp_path)


def test_executor_unknown_format_raises():
    with pytest.raises(NotImplementedError, match="not yet supported"):
        run_executor.execute_run_sync(
            model_bytes=HW_MODEL.read_bytes(),
            pack_bytes=HW_INPUTS.read_bytes(),
            output_template_bytes=HW_REPORT.read_bytes(),
            output_template_format="pdf",
        )


def test_executor_handles_pack_missing_input_tab():
    """Pack file contract violation should bubble up as ValueError."""
    # Build a pack that doesn't have the I_Numbers tab the Model expects
    wb = openpyxl.Workbook()
    del wb[wb.sheetnames[0]]
    wb.create_sheet("I_Other")
    buf = io.BytesIO()
    wb.save(buf)
    bad_pack = buf.getvalue()

    with pytest.raises(ValueError, match="missing required input tabs"):
        run_executor.execute_run_sync(
            model_bytes=HW_MODEL.read_bytes(),
            pack_bytes=bad_pack,
            output_template_bytes=HW_REPORT.read_bytes(),
            output_template_format="xlsx",
        )
