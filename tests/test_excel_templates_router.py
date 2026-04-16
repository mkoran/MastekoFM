"""Tests for Excel Templates router (upload, list, get, replace, delete)."""
from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl
from fastapi.testclient import TestClient

from backend.app.main import app

client = TestClient(app)
AUTH = {"Authorization": "Bearer dev-test@example.com"}
PATCH_DB = "backend.app.routers.excel_templates.get_firestore_client"
PATCH_STORAGE_UPLOAD = "backend.app.routers.excel_templates.storage_service.upload_xlsx"
PATCH_STORAGE_DELETE = "backend.app.routers.excel_templates.storage_service.delete_blob"
PATCH_STORAGE_DOWNLOAD = "backend.app.routers.excel_templates.storage_service.download_xlsx"

FIXTURE = Path(__file__).parent / "fixtures" / "campus_adele.xlsx"


def _tiny_template_bytes() -> bytes:
    """Minimal valid template for router tests that don't need the full fixture."""
    wb = openpyxl.Workbook()
    del wb[wb.sheetnames[0]]
    wb.create_sheet("I_Inputs")
    wb["I_Inputs"]["A1"] = "val"
    wb.create_sheet("O_Outputs")
    wb.create_sheet("Calc")
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@patch(PATCH_STORAGE_UPLOAD, return_value="https://example/path")
@patch(PATCH_DB)
def test_upload_excel_template_classifies_tabs(mock_db, mock_upload):
    mock_client = MagicMock()
    mock_doc_ref = MagicMock()
    mock_doc_ref.id = "tpl-1"
    mock_client.collection.return_value.document.return_value = mock_doc_ref
    mock_db.return_value = mock_client

    file_bytes = _tiny_template_bytes()
    resp = client.post(
        "/api/excel-templates",
        headers=AUTH,
        files={"file": ("toy.xlsx", file_bytes,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"name": "Toy", "code_name": "toy"},
    )
    assert resp.status_code == 201, resp.text
    body = resp.json()
    assert body["input_tabs"] == ["I_Inputs"]
    assert body["output_tabs"] == ["O_Outputs"]
    assert body["calc_tabs"] == ["Calc"]
    assert body["version"] == 1
    assert mock_upload.called


@patch(PATCH_DB)
def test_upload_rejects_file_without_input_tab(mock_db):
    mock_client = MagicMock()
    mock_doc_ref = MagicMock()
    mock_doc_ref.id = "tpl-x"
    mock_client.collection.return_value.document.return_value = mock_doc_ref
    mock_db.return_value = mock_client

    wb = openpyxl.Workbook()
    del wb[wb.sheetnames[0]]
    wb.create_sheet("Calc")
    buf = BytesIO()
    wb.save(buf)

    resp = client.post(
        "/api/excel-templates",
        headers=AUTH,
        files={"file": ("bad.xlsx", buf.getvalue(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"name": "Bad"},
    )
    assert resp.status_code == 400
    assert "I_" in resp.json()["detail"]


@patch(PATCH_DB)
def test_list_excel_templates(mock_db):
    mock_client = MagicMock()
    mock_db.return_value = mock_client
    now = datetime.now(UTC)
    mock_doc = MagicMock()
    mock_doc.id = "tpl-1"
    mock_doc.to_dict.return_value = {
        "name": "Campus Adele", "code_name": "campus_adele", "version": 1,
        "input_tabs": ["I_Inputs"], "output_tabs": ["O_Out"], "calc_tabs": ["C"],
        "created_at": now, "updated_at": now,
    }
    mock_client.collection.return_value.stream.return_value = [mock_doc]

    resp = client.get("/api/excel-templates", headers=AUTH)
    assert resp.status_code == 200
    data = resp.json()
    assert len(data) == 1
    assert data[0]["input_tab_count"] == 1


def test_upload_requires_auth():
    resp = client.post(
        "/api/excel-templates",
        files={"file": ("x.xlsx", b"fake",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"name": "x"},
    )
    assert resp.status_code == 401
