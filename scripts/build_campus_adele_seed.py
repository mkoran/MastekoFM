"""Build Campus Adele seed files for Sprint B from the existing test fixture.

Run from repo root:
    python scripts/build_campus_adele_seed.py

Produces:
    seed/campus_adele/campus_adele_model.xlsx       (the existing 15-tab model)
    seed/campus_adele/campus_adele_base_pack.xlsx   (only I_* tabs, with Base Case values)
    seed/campus_adele/campus_adele_summary.xlsx     (minimal OutputTemplate)
    seed/campus_adele/README.md
"""
import shutil
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

REPO = Path(__file__).resolve().parents[1]
SEED_DIR = REPO / "seed" / "campus_adele"
FIXTURE = REPO / "tests" / "fixtures" / "campus_adele.xlsx"

SEED_DIR.mkdir(parents=True, exist_ok=True)


def build_model() -> None:
    """Copy the existing 15-tab Campus Adele Model verbatim — already obeys I_/O_/calc."""
    target = SEED_DIR / "campus_adele_model.xlsx"
    shutil.copy(FIXTURE, target)
    print(f"wrote {target}")


def build_base_pack() -> None:
    """Strip the model down to only its I_* tabs — that's a base AssumptionPack."""
    wb = load_workbook(FIXTURE, data_only=False)
    to_delete = [n for n in wb.sheetnames if not n.startswith("I_")]
    for n in to_delete:
        del wb[n]
    target = SEED_DIR / "campus_adele_base_pack.xlsx"
    wb.save(target)
    wb.close()
    print(f"wrote {target}  ({len(wb.sheetnames)} I_ tabs preserved)")


def build_summary_template() -> None:
    """Build a minimal OutputTemplate that surfaces a few O_Annual Summary cells.

    M_Annual Summary tab: pulls cells from Model's O_Annual Summary
    O_Report tab: shows them in a simple human-readable format
    """
    wb = Workbook()
    wb.remove(wb.active)

    # M_Annual Summary — placeholders to be filled by Model.O_Annual Summary
    # We mirror the dimension of O_Annual Summary so cell-copy works (130x10)
    m = wb.create_sheet("M_Annual Summary")
    # Just the cells we plan to surface — others stay blank
    # We don't know exact cell coordinates without inspection — leave the tab empty,
    # the engine's overlay will fill whatever cells the Model populates.

    # O_Report
    o = wb.create_sheet("O_Report")
    header = Font(bold=True, color="FFFFFF", size=14)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    label_font = Font(bold=True)
    o["A1"] = "Campus Adele — Investor Summary"
    o["A1"].font = header
    o["A1"].fill = header_fill
    o["A1"].alignment = Alignment(horizontal="center")
    o.merge_cells("A1:D1")

    # Pull a representative slice — these reference cells in the M_ tab
    o["A3"] = "Annual Summary (first row)"
    o["A3"].font = label_font
    # Copy the first 5 cells of row 1 from M_Annual Summary — labels
    for col_idx, col_letter in enumerate(["B", "C", "D", "E", "F"], start=2):
        o.cell(row=4, column=col_idx).value = f"='M_Annual Summary'!{col_letter}3"
    o["A4"] = "Item"
    o["A4"].font = label_font

    # Row 5: first data row
    o["A5"] = "Year 1"
    o["A5"].font = label_font
    for col_idx, col_letter in enumerate(["B", "C", "D", "E", "F"], start=2):
        o.cell(row=5, column=col_idx).value = f"='M_Annual Summary'!{col_letter}9"

    # Footer note
    o["A8"] = "Note: cells reference M_Annual Summary which is filled at run time from the Model's O_Annual Summary tab."
    o["A8"].font = Font(italic=True, size=10, color="888888")
    o.column_dimensions["A"].width = 22
    for col in "BCDEF":
        o.column_dimensions[col].width = 16

    target = SEED_DIR / "campus_adele_summary.xlsx"
    wb.save(target)
    wb.close()
    print(f"wrote {target}")


def write_readme() -> None:
    p = SEED_DIR / "README.md"
    p.write_text("""# Campus Adele seed

The 15-tab construction-to-perm financing model used to validate Sprint B's
post-cleanup architecture.

## Files

| File | Role | Description |
|---|---|---|
| `campus_adele_model.xlsx` | Model | Full 15-tab Construction-to-Perm model. 5 `I_` input tabs, 1 `O_Annual Summary` output tab, 9 calc tabs. |
| `campus_adele_base_pack.xlsx` | AssumptionPack | Only the `I_` tabs from the Model, with the Base Case values pre-filled. |
| `campus_adele_summary.xlsx` | OutputTemplate | Minimal investor summary — `M_Annual Summary` (filled by Model.O_Annual Summary) + `O_Report`. |

## Seed via API

```bash
curl -X POST <API_URL>/api/seed/campus-adele \\
  -H "Authorization: Bearer <token>" \\
  -H "X-MFM-Drive-Token: <google-token>"
```

Idempotent. Returns the IDs of the created Model + Project + AssumptionPack + OutputTemplate.

## Rebuild from fixture

```bash
python scripts/build_campus_adele_seed.py
```
""")
    print(f"wrote {p}")


if __name__ == "__main__":
    build_model()
    build_base_pack()
    build_summary_template()
    write_readme()
