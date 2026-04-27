"""One-shot script to build the Hello World seed .xlsx files.

Run from repo root:
    python scripts/build_helloworld_seed.py

Produces:
    seed/helloworld/helloworld_model.xlsx
    seed/helloworld/helloworld_inputs.xlsx
    seed/helloworld/helloworld_report.xlsx

These are committed to the repo as canonical fixtures for both /api/seed/helloworld
and tests/fixtures/.
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

SEED_DIR = Path(__file__).resolve().parents[1] / "seed" / "helloworld"
SEED_DIR.mkdir(parents=True, exist_ok=True)

HEADER_FONT = Font(bold=True, size=12)
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_TEXT = Font(bold=True, color="FFFFFF", size=14)


def build_model() -> None:
    """helloworld_model.xlsx — the Model.

    Tabs: I_Numbers, Calc, O_Results
    """
    wb = Workbook()
    wb.remove(wb.active)

    # I_Numbers
    ws_in = wb.create_sheet("I_Numbers")
    ws_in["A1"] = "a"; ws_in["B1"] = 2
    ws_in["A2"] = "b"; ws_in["B2"] = 3
    ws_in["A1"].font = HEADER_FONT
    ws_in["A2"].font = HEADER_FONT
    ws_in["B1"].fill = INPUT_FILL
    ws_in["B2"].fill = INPUT_FILL
    ws_in.column_dimensions["A"].width = 10
    ws_in.column_dimensions["B"].width = 10

    # Calc
    ws_calc = wb.create_sheet("Calc")
    ws_calc["A1"] = "=I_Numbers!B1+I_Numbers!B2"
    ws_calc["A2"] = "=I_Numbers!B1*I_Numbers!B2"

    # O_Results
    ws_out = wb.create_sheet("O_Results")
    ws_out["A1"] = "sum"; ws_out["B1"] = "=Calc!A1"
    ws_out["A2"] = "product"; ws_out["B2"] = "=Calc!A2"
    ws_out["A1"].font = HEADER_FONT
    ws_out["A2"].font = HEADER_FONT
    ws_out.column_dimensions["A"].width = 12
    ws_out.column_dimensions["B"].width = 12

    path = SEED_DIR / "helloworld_model.xlsx"
    wb.save(path)
    print(f"wrote {path}")


def build_inputs() -> None:
    """helloworld_inputs.xlsx — the AssumptionPack (only I_ tabs)."""
    wb = Workbook()
    wb.remove(wb.active)

    ws_in = wb.create_sheet("I_Numbers")
    ws_in["A1"] = "a"; ws_in["B1"] = 5
    ws_in["A2"] = "b"; ws_in["B2"] = 7
    ws_in["A1"].font = HEADER_FONT
    ws_in["A2"].font = HEADER_FONT
    ws_in["B1"].fill = INPUT_FILL
    ws_in["B2"].fill = INPUT_FILL
    ws_in.column_dimensions["A"].width = 10
    ws_in.column_dimensions["B"].width = 10

    path = SEED_DIR / "helloworld_inputs.xlsx"
    wb.save(path)
    print(f"wrote {path}")


def build_report() -> None:
    """helloworld_report.xlsx — the OutputTemplate.

    Tabs: M_Results (placeholders to be filled by Model O_Results), O_Report (the artifact)
    """
    wb = Workbook()
    wb.remove(wb.active)

    # M_Results — placeholders
    ws_m = wb.create_sheet("M_Results")
    ws_m["A1"] = "sum"; ws_m["B1"] = 0
    ws_m["A2"] = "product"; ws_m["B2"] = 0
    ws_m["A1"].font = HEADER_FONT
    ws_m["A2"].font = HEADER_FONT
    ws_m.column_dimensions["A"].width = 12
    ws_m.column_dimensions["B"].width = 12

    # O_Report — the user-facing artifact
    ws_r = wb.create_sheet("O_Report")
    ws_r["A1"] = "Hello World Report"
    ws_r["A1"].font = HEADER_TEXT
    ws_r["A1"].fill = HEADER_FILL
    ws_r["A1"].alignment = Alignment(horizontal="center")
    ws_r.merge_cells("A1:B1")

    ws_r["A3"] = "Sum:"; ws_r["B3"] = "=M_Results!B1"
    ws_r["A4"] = "Product:"; ws_r["B4"] = "=M_Results!B2"
    ws_r["A5"] = "Total:"; ws_r["B5"] = "=M_Results!B1+M_Results!B2"

    for cell in ("A3", "A4", "A5"):
        ws_r[cell].font = Font(bold=True)
    ws_r.column_dimensions["A"].width = 14
    ws_r.column_dimensions["B"].width = 14

    path = SEED_DIR / "helloworld_report.xlsx"
    wb.save(path)
    print(f"wrote {path}")


if __name__ == "__main__":
    build_model()
    build_inputs()
    build_report()
    print("\nHello World seed files built. Verify with:")
    print("  python -c \"from openpyxl import load_workbook; "
          "wb = load_workbook('seed/helloworld/helloworld_model.xlsx'); print(wb.sheetnames)\"")
